from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    in_path = repo_root / "System Test.xlsx"
    out_path = repo_root / "System Test.xlsx"
    sheet_to_keep = "Club Fund"

    import sys

    if len(sys.argv) >= 2:
        in_path = Path(sys.argv[1])
    if len(sys.argv) >= 3:
        sheet_to_keep = sys.argv[2]
    if len(sys.argv) >= 4:
        out_path = Path(sys.argv[3])

    wb = load_workbook(in_path, data_only=False)
    if sheet_to_keep not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet_to_keep}' not found.")

    drop = [s for s in wb.sheetnames if s != sheet_to_keep]
    for s in drop:
        wb.remove(wb[s])

    wb.active = 0
    wb.save(out_path)
    print(f"kept=1 dropped={len(drop)} out={out_path}")


if __name__ == "__main__":
    main()

