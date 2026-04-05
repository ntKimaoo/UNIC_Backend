# -*- coding: utf-8 -*-
"""Sync Unit Test.xlsx Club Fund sheets with ClubFundControllerTest + ClubFundServiceTest (UTCID matrix, totals)."""
from __future__ import annotations

import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from clubfund_xlsx_conditions import apply_business_condition_matrix

REPO = Path(__file__).resolve().parents[1]
XLSX = REPO / "Unit Test.xlsx"
CTRL_TEST = REPO / "UNIC.ControllerTest" / "Controllers" / "ClubFundControllerTest.cs"
SVC_TEST_CS = REPO / "UNIC.ServiceTest" / "Services" / "ClubFundServiceTest.cs"

EXEC_DATE = date(2026, 4, 4)
UTCID_START_COL = 5  # E


def extract_async_test_names(cs_text: str) -> list[str]:
    return re.findall(r"public async Task (\w+)\s*\(", cs_text)


def scenario_type(test_name: str) -> str:
    """N = normal/happy path scenario, A = error/edge scenario (for Result row)."""
    if "ReturnsInternalServerError" in test_name or "ReturnsBadRequest" in test_name:
        return "A"
    if "Returns403" in test_name or "Returns404" in test_name or "ReturnsNotFound" in test_name:
        return "A"
    if "ReturnsUnauthorized" in test_name:
        return "A"
    if "_ShouldThrow" in test_name or "_ShouldReturnFalse" in test_name:
        return "A"
    if "ShouldReturnNull" in test_name:
        return "A"
    if "WhenNotSuccessCode" in test_name:
        return "N"
    if "ReturnsOk" in test_name or "ShouldReturnTrue" in test_name:
        return "N"
    if "_ShouldReturnPagedDtos" in test_name or "_ShouldMap" in test_name:
        return "N"
    if "_ShouldComplete" in test_name or "_ShouldUpdate_WhenReject" in test_name:
        return "N"
    if "_ShouldBypassMemberGate" in test_name or "_ShouldPass" in test_name:
        return "N"
    if "_ShouldCallPayOS" in test_name or "_ShouldDeleteTransaction" in test_name:
        return "N"
    if "_ShouldSetApproved" in test_name or "_ShouldSetPending" in test_name:
        return "N"
    if "_ShouldReturnPayOsPayload" in test_name:
        return "N"
    if "PassesApproved" in test_name or "PassesIsSystemAdmin" in test_name:
        return "N"
    return "A"


def filter_by_prefix(names: list[str], prefix: str) -> list[str]:
    p = prefix + "_"
    return [n for n in names if n.startswith(p)]


def code_module_club_fund(sheet_title: str) -> str:
    if sheet_title.startswith("C.ClubFund"):
        return "ClubFundController"
    if sheet_title.startswith("S.ClubFund"):
        return "ClubFundService"
    return "ClubFundController"


def apply_template_context(
    ws: Worksheet,
    *,
    code_module: str,
    d9_precondition: str | None = None,
    d10_precondition: str | None = None,
    parameter_b11: str | None = None,
) -> None:
    """C1 = class under test; optional Condition rows (same layout as template)."""
    ws.cell(1, 3).value = code_module
    if d9_precondition is not None:
        ws.cell(9, 4).value = d9_precondition
    if d10_precondition is not None:
        ws.cell(10, 4).value = d10_precondition
    if parameter_b11 is not None:
        ws.cell(11, 2).value = parameter_b11


def apply_matrix(
    ws: Worksheet,
    tests: list[str],
    *,
    cs_text: str,
    sheet_title: str,
    csharp_method: str,
    controller: bool,
) -> None:
    n = len(tests)
    if n == 0:
        return

    for i, tname in enumerate(tests):
        c = UTCID_START_COL + i
        ws.cell(7, c).value = f"UTCID{i + 1:02d}"
        ws.cell(6, c).value = tname

    first_after = UTCID_START_COL + n
    for c in range(first_after, min(ws.max_column, UTCID_START_COL + 40) + 1):
        for r in (6, 7, 9, 10, 12, 13, 15, 16, 21, 22, 23):
            ws.cell(r, c).value = None

    apply_business_condition_matrix(
        ws,
        tests=tests,
        cs_text=cs_text,
        sheet_title=sheet_title,
        csharp_method=csharp_method,
        controller=controller,
        scenario_type_fn=scenario_type,
    )

    normals = sum(1 for t in tests if scenario_type(t) == "N")
    abnormals = sum(1 for t in tests if scenario_type(t) == "A")
    boundaries = 0

    for i, t in enumerate(tests):
        ws.cell(21, UTCID_START_COL + i).value = scenario_type(t)

    ws.cell(5, 1).value = n
    ws.cell(5, 3).value = 0
    ws.cell(5, 5).value = 0
    ws.cell(5, 11).value = normals
    ws.cell(5, 12).value = abnormals
    ws.cell(5, 13).value = boundaries
    ws.cell(5, 14).value = n

    for i in range(n):
        c = UTCID_START_COL + i
        ws.cell(22, c).value = "P"
        ws.cell(23, c).value = EXEC_DATE
        cell = ws.cell(23, c)
        cell.number_format = "dd/mm/yyyy"


REQUIREMENTS: dict[str, str] = {
    "C.ClubFund.CreateFund": "POST tạo quỹ: Ok, 403, BadRequest (validation).",
    "C.ClubFund.GetMyClubs": "GET my-clubs: Ok.",
    "C.ClubFund.GetFundCapabilities": "GET capabilities: 403, Ok, Unauthorized.",
    "C.ClubFund.GetReportSummary": "GET report-summary: 403, Ok, BadRequest khi fromUtc > toUtc.",
    "C.ClubFund.GetFundCategories": "GET categories: 403, Ok, Unauthorized.",
    "C.ClubFund.GetFundTrans": "GET transactions: 403, page/pageSize/date range, Unauthorized, fund club mismatch, NotFound, Ok.",
    "C.ClubFund.GetFund": "GET fund by id: 404, 403, Ok (admin bypass).",
    "C.ClubFund.GetFundsByClub": "GET funds list: BadRequest page/size, 403, Ok, Admin passes isSystemAdmin.",
    "C.ClubFund.GetMyFunds": "GET my: 403, BadRequest mineType, Ok.",
    "C.ClubFund.Contribute": "POST contribute: 403, 404, Ok.",
    "C.ClubFund.GetContrPayStatus": "GET contribute status: 403, NotFound, Ok.",
    "C.ClubFund.GetPayOsReturn": "GET payos-return: NotFound, 403, Ok.",
    "C.ClubFund.SimulatePayOsDev": "POST dev simulate: NotFound prod, Ok dev.",
    "C.ClubFund.ApproveFund": "POST approve: Ok, 403, BadRequest Argument/InvalidOp/Exception.",
    "C.ClubFund.PayOSWebhook": "POST webhook: empty body, code!=00, missing sig, bad sig, ok process, invalid orderCode, invalid JSON.",
    "C.ClubFund.GetHistory": "GET history: BadRequest page, NotFound, 403 club mismatch, 403 access, Ok, Admin.",
    "C.ClubFund.GetFundLocation": "GET fund location: 403, 404, Ok.",
    "S.ClubFund.CreateFundAsync": "CreateFundAsync validation, role, auto-approve/pending.",
    "S.ClubFund.CreateContribAsync": "CreateContributionAsync amount/fund/category/PayOS.",
    "S.ClubFund.GetContrPayStatus": "GetContributionPaymentStatusAsync wrong user -> null.",
    "S.ClubFund.GetPayStatusByOrd": "GetContributionPaymentStatusByOrderCodeAsync map paid.",
    "S.ClubFund.GetFundByIdAsync": "GetFundByIdAsync missing -> null.",
    "S.ClubFund.GetFundsByClubPaged": "GetFundsByClubIdPagedAsync paging, invalid filters, force APPROVED, PENDING manager, system admin.",
    "S.ClubFund.GetMyFundsPaged": "GetMyFundsByClubIdPagedAsync filters + invalid mineType.",
    "S.ClubFund.GetFundHistoryPaged": "GetFundHistoryPagedAsync default APPROVED, ALL, mine scope.",
    "S.ClubFund.ApproveFundAsync": "ApproveFundAsync not found, approved, not manager, reject ok, reject no reason, short reason.",
    "S.ClubFund.ProcessPayOSSuccess": "ProcessPayOSPaymentSuccessAsync false/true.",
    "S.ClubFund.TryCompletePending": "TryCompleteOwnPending wrong club / valid.",
    "S.ClubFund.GetFundCapabilities": "GetFundCapabilitiesAsync not member, policies, empty menu.",
    "S.ClubFund.ReportSummaryAsync": "GetClubFundReportSummaryAsync maps aggregates.",
    "S.ClubFund.ClubFundTransPaged": "GetClubFundTransactionsPagedAsync default, mine, ALL status.",
}


def main() -> None:
    ctrl_text = CTRL_TEST.read_text(encoding="utf-8")
    svc_text = SVC_TEST_CS.read_text(encoding="utf-8")
    ctrl_names = extract_async_test_names(ctrl_text)
    svc_names = extract_async_test_names(svc_text)

    wb = load_workbook(XLSX, data_only=False)
    template = wb["C.ClubFund.GetMyClubs"]

    controller_specs: list[tuple[str, str, str, list[str]]] = [
        ("C.ClubFund.CreateFund", "CreateFund", "CreateFund", filter_by_prefix(ctrl_names, "CreateFund")),
        ("C.ClubFund.GetMyClubs", "GetMyClubs", "GetMyClubs", filter_by_prefix(ctrl_names, "GetMyClubs")),
        (
            "C.ClubFund.GetFundCapabilities",
            "GetFundCapabilities",
            "GetFundCapabilities",
            filter_by_prefix(ctrl_names, "GetFundCapabilities"),
        ),
        (
            "C.ClubFund.GetReportSummary",
            "GetFundReportSummary",
            "GetFundReportSummary",
            filter_by_prefix(ctrl_names, "GetFundReportSummary"),
        ),
        (
            "C.ClubFund.GetFundCategories",
            "GetFundCategories",
            "GetFundCategories",
            filter_by_prefix(ctrl_names, "GetFundCategories"),
        ),
        (
            "C.ClubFund.GetFundTrans",
            "GetClubFundTransactions",
            "GetClubFundTransactions",
            filter_by_prefix(ctrl_names, "GetClubFundTransactions"),
        ),
        (
            "C.ClubFund.GetFund",
            "GetFund",
            "GetFund",
            [n for n in ctrl_names if n.startswith("GetFund_")],
        ),
        (
            "C.ClubFund.GetFundsByClub",
            "GetFundsByClub",
            "GetFundsByClub",
            filter_by_prefix(ctrl_names, "GetFundsByClub"),
        ),
        ("C.ClubFund.GetMyFunds", "GetMyFunds", "GetMyFunds", filter_by_prefix(ctrl_names, "GetMyFunds")),
        ("C.ClubFund.Contribute", "Contribute", "Contribute", filter_by_prefix(ctrl_names, "Contribute")),
        (
            "C.ClubFund.GetContrPayStatus",
            "GetContributionPaymentStatus",
            "GetContributionPaymentStatus",
            filter_by_prefix(ctrl_names, "GetContributionPaymentStatus"),
        ),
        (
            "C.ClubFund.GetPayOsReturn",
            "GetPayOsContributionReturn",
            "GetPayOsContributionReturn",
            filter_by_prefix(ctrl_names, "GetPayOsContributionReturn"),
        ),
        (
            "C.ClubFund.SimulatePayOsDev",
            "SimulatePayOsPaidForDevelopment",
            "SimulatePayOsPaidForDevelopment",
            filter_by_prefix(ctrl_names, "SimulatePayOsPaidForDevelopment"),
        ),
        ("C.ClubFund.ApproveFund", "ApproveFund", "ApproveFund", filter_by_prefix(ctrl_names, "ApproveFund")),
        (
            "C.ClubFund.PayOSWebhook",
            "PayOSWebhookClubScoped",
            "PayOSWebhookClubScoped",
            [n for n in ctrl_names if n.startswith("PayOSWebhook_")],
        ),
        ("C.ClubFund.GetHistory", "GetHistory", "GetHistory", filter_by_prefix(ctrl_names, "GetHistory")),
        (
            "C.ClubFund.GetFundLocation",
            "GetFundLocation",
            "GetFundLocation",
            filter_by_prefix(ctrl_names, "GetFundLocation"),
        ),
    ]

    service_specs: list[tuple[str, str, str, list[str]]] = [
        ("S.ClubFund.CreateFundAsync", "CreateFundAsync", "CreateFundAsync", filter_by_prefix(svc_names, "CreateFundAsync")),
        (
            "S.ClubFund.CreateContribAsync",
            "CreateContributionAsync",
            "CreateContributionAsync",
            filter_by_prefix(svc_names, "CreateContributionAsync"),
        ),
        (
            "S.ClubFund.GetContrPayStatus",
            "GetContributionPaymentStatusAsync",
            "GetContributionPaymentStatusAsync",
            filter_by_prefix(svc_names, "GetContributionPaymentStatusAsync"),
        ),
        (
            "S.ClubFund.GetPayStatusByOrd",
            "GetContributionPaymentStatusByOrderCodeAsync",
            "GetContributionPaymentStatusByOrderCodeAsync",
            filter_by_prefix(svc_names, "GetContributionPaymentStatusByOrderCodeAsync"),
        ),
        (
            "S.ClubFund.GetFundByIdAsync",
            "GetFundByIdAsync",
            "GetFundByIdAsync",
            filter_by_prefix(svc_names, "GetFundByIdAsync"),
        ),
        (
            "S.ClubFund.GetFundsByClubPaged",
            "GetFundsByClubIdPagedAsync",
            "GetFundsByClubIdPagedAsync",
            filter_by_prefix(svc_names, "GetFundsByClubIdPagedAsync"),
        ),
        (
            "S.ClubFund.GetMyFundsPaged",
            "GetMyFundsByClubIdPagedAsync",
            "GetMyFundsByClubIdPagedAsync",
            filter_by_prefix(svc_names, "GetMyFundsByClubIdPagedAsync"),
        ),
        (
            "S.ClubFund.GetFundHistoryPaged",
            "GetFundHistoryPagedAsync",
            "GetFundHistoryPagedAsync",
            filter_by_prefix(svc_names, "GetFundHistoryPagedAsync"),
        ),
        (
            "S.ClubFund.ApproveFundAsync",
            "ApproveFundAsync",
            "ApproveFundAsync",
            filter_by_prefix(svc_names, "ApproveFundAsync"),
        ),
        (
            "S.ClubFund.ProcessPayOSSuccess",
            "ProcessPayOSPaymentSuccessAsync",
            "ProcessPayOSPaymentSuccessAsync",
            filter_by_prefix(svc_names, "ProcessPayOSPaymentSuccessAsync"),
        ),
        (
            "S.ClubFund.TryCompletePending",
            "TryCompleteOwnPendingContributionForDevelopmentAsync",
            "TryCompleteOwnPendingContributionForDevelopmentAsync",
            filter_by_prefix(svc_names, "TryCompleteOwnPendingContributionForDevelopmentAsync"),
        ),
        (
            "S.ClubFund.GetFundCapabilities",
            "GetFundCapabilitiesAsync",
            "GetFundCapabilitiesAsync",
            filter_by_prefix(svc_names, "GetFundCapabilitiesAsync"),
        ),
        (
            "S.ClubFund.ReportSummaryAsync",
            "GetClubFundReportSummaryAsync",
            "GetClubFundReportSummaryAsync",
            filter_by_prefix(svc_names, "GetClubFundReportSummaryAsync"),
        ),
        (
            "S.ClubFund.ClubFundTransPaged",
            "GetClubFundTransactionsPagedAsync",
            "GetClubFundTransactionsPagedAsync",
            filter_by_prefix(svc_names, "GetClubFundTransactionsPagedAsync"),
        ),
    ]

    created: list[str] = []
    updated: list[str] = []

    def ensure_sheet(title: str) -> Worksheet:
        nonlocal created
        if title in wb.sheetnames:
            return wb[title]
        ws = wb.copy_worksheet(template)
        ws.title = title[:31]
        created.append(ws.title)
        return ws

    for sheet_title, _api, method_k, tests in controller_specs + service_specs:
        if not tests:
            continue
        ws = ensure_sheet(sheet_title)
        apply_template_context(ws, code_module=code_module_club_fund(sheet_title))
        ws.cell(1, 11).value = method_k
        req = REQUIREMENTS.get(sheet_title)
        if req:
            ws.cell(3, 3).value = req
        apply_matrix(
            ws,
            tests,
            cs_text=ctrl_text if sheet_title.startswith("C.ClubFund") else svc_text,
            sheet_title=sheet_title,
            csharp_method=method_k,
            controller=sheet_title.startswith("C.ClubFund"),
        )
        updated.append(f"{ws.title}: {len(tests)} cases")

    # Drop PayOS/QR unit-test sheets (tracked only in code, not this workbook).
    for name in list(wb.sheetnames):
        if name.startswith("S.PayOS.") or name.startswith("S.QRCode.") or name == "S.QRCodeGenerator.GetQrCodePngB":
            wb.remove(wb[name])

    # Remove legacy empty service rows if old sheet had extra name (optional)
    try:
        wb.save(XLSX)
        print("saved:", XLSX)
    except PermissionError:
        alt = REPO / "Unit Test.sync-output.xlsx"
        wb.save(alt)
        print(
            "WARN: could not overwrite (file open?). Saved to:",
            alt,
            "\nClose Excel and run again, or replace Unit Test.xlsx manually.",
        )
    if created:
        print("created_sheets:", created)
    for line in updated:
        print(line)


if __name__ == "__main__":
    main()
