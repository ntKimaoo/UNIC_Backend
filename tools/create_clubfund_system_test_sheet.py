from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


@dataclass(frozen=True)
class TC:
    tc_id: str
    desc: str
    steps: str
    expected: str
    pre: str


def build_cases(prefix: str = "CF_3003") -> list[tuple[str, list[TC]]]:
    """
    Returns scenarios -> list of test cases.
    IDs follow pattern CF_DDMM_scenario.tc (similar to MR_1403_1.01).
    """
    def tid(scn: int, idx: int) -> str:
        return f"{prefix}_{scn}.{idx:02d}"

    scenarios: list[tuple[str, list[TC]]] = []

    scenarios.append(
        (
            "Scenario 1- Create Fund",
            [
                TC(
                    tid(1, 1),
                    "Tạo quỹ thành công khi user là Manager cấp cao nhất (status = APPROVED)",
                    "1. Đăng nhập bằng user là Manager level 1 của CLB.\n"
                    "2. Vào màn hình tạo quỹ.\n"
                    "3. Nhập dữ liệu hợp lệ (tên, số tiền, hạn nếu có).\n"
                    "4. Submit.",
                    "Tạo quỹ thành công.\n"
                    "Quỹ được tạo với status = APPROVED.\n"
                    "Thông báo thành công hiển thị, redirect về chi tiết quỹ hoặc danh sách quỹ.",
                    "User thuộc CLB và có role Manager level 1.\nCLB tồn tại và user đang ACTIVE.",
                ),
                TC(
                    tid(1, 2),
                    "Tạo quỹ thành công khi user là Vice Manager (status = PENDING)",
                    "1. Đăng nhập bằng user là Vice Manager level 2.\n"
                    "2. Tạo quỹ với dữ liệu hợp lệ.\n"
                    "3. Submit.",
                    "Tạo quỹ thành công.\nQuỹ được tạo với status = PENDING.\nHiển thị thông báo chờ duyệt.",
                    "User thuộc CLB và có role Vice Manager level 2 (ACTIVE).",
                ),
                TC(
                    tid(1, 3),
                    "Không tạo được quỹ khi tên quỹ trống/space",
                    "1. Đăng nhập user đủ quyền.\n2. Vào tạo quỹ.\n3. Để trống FundName hoặc nhập toàn space.\n4. Submit.",
                    "Hiển thị lỗi validation rõ ràng cho trường tên quỹ.\nKhông tạo quỹ.",
                    "User đủ quyền tạo quỹ.",
                ),
                TC(
                    tid(1, 4),
                    "Không tạo được quỹ khi initial amount âm",
                    "1. Đăng nhập user đủ quyền.\n2. Nhập InitialAmount = -1.\n3. Submit.",
                    "Hiển thị lỗi validation.\nKhông tạo quỹ.",
                    "User đủ quyền tạo quỹ.",
                ),
                TC(
                    tid(1, 5),
                    "Không tạo được quỹ khi ExpiresAt là ngày trong quá khứ (nếu có hạn)",
                    "1. Đăng nhập user đủ quyền.\n2. Chọn hạn đóng quỹ < hôm nay.\n3. Submit.",
                    "Hiển thị lỗi validation.\nKhông tạo quỹ.",
                    "User đủ quyền tạo quỹ.",
                ),
                TC(
                    tid(1, 6),
                    "Không tạo được quỹ khi user không phải member của CLB",
                    "1. Đăng nhập user không thuộc CLB.\n2. Truy cập tạo quỹ theo clubId.\n3. Submit.",
                    "Bị chặn quyền (403 hoặc thông báo không có quyền).\nKhông tạo quỹ.",
                    "User đăng nhập nhưng không thuộc CLB.",
                ),
                TC(
                    tid(1, 7),
                    "Không tạo được quỹ khi member inactive/LEFT",
                    "1. Đăng nhập user từng là member nhưng status != ACTIVE.\n2. Thử tạo quỹ.\n3. Submit.",
                    "Bị chặn quyền.\nKhông tạo quỹ.",
                    "User thuộc CLB nhưng trạng thái membership không ACTIVE.",
                ),
                TC(
                    tid(1, 8),
                    "Không tạo được quỹ khi user là member thường (không phải quản lý)",
                    "1. Đăng nhập user role member thường.\n2. Thử tạo quỹ.\n3. Submit.",
                    "Bị chặn quyền.\nKhông tạo quỹ.",
                    "User ACTIVE nhưng role không đủ quyền tạo quỹ.",
                ),
            ],
        )
    )

    scenarios.append(
        (
            "Scenario 2- Contribute (Create contribution)",
            [
                TC(
                    tid(2, 1),
                    "Đóng góp thành công và nhận link PayOS khi số tiền hợp lệ",
                    "1. Đăng nhập user là member của CLB.\n"
                    "2. Mở chi tiết quỹ status = APPROVED.\n"
                    "3. Nhập Amount >= min.\n"
                    "4. Submit đóng góp.",
                    "Tạo contribution thành công.\nNhận được thông tin thanh toán PayOS (paymentUrl/qr/orderCode).\nTrạng thái đóng góp = PENDING (chờ thanh toán).",
                    "Quỹ tồn tại, status = APPROVED, chưa hết hạn.\nPayOS sandbox configured.",
                ),
                TC(
                    tid(2, 2),
                    "Không tạo đóng góp khi Amount = 0",
                    "1. Đăng nhập.\n2. Chọn quỹ.\n3. Nhập Amount = 0.\n4. Submit.",
                    "Hiển thị lỗi validation số tiền.\nKhông tạo contribution.",
                    "Quỹ tồn tại.",
                ),
                TC(
                    tid(2, 3),
                    "Không tạo đóng góp khi Amount < minimum",
                    "1. Đăng nhập.\n2. Nhập Amount nhỏ hơn mức tối thiểu.\n3. Submit.",
                    "Hiển thị lỗi validation.\nKhông tạo contribution.",
                    "Quỹ có rule min amount.",
                ),
                TC(
                    tid(2, 4),
                    "Không tạo đóng góp khi quỹ không tồn tại",
                    "1. Đăng nhập.\n2. Truy cập đóng góp với fundId không tồn tại.\n3. Submit.",
                    "Hiển thị lỗi NotFound hoặc thông báo quỹ không tồn tại.\nKhông tạo contribution.",
                    "User đăng nhập.",
                ),
                TC(
                    tid(2, 5),
                    "Không tạo đóng góp khi quỹ chưa được duyệt (PENDING)",
                    "1. Đăng nhập.\n2. Chọn quỹ status = PENDING.\n3. Thử đóng góp.",
                    "Bị chặn thao tác.\nThông báo quỹ chưa được duyệt.\nKhông tạo contribution.",
                    "Có quỹ PENDING.",
                ),
                TC(
                    tid(2, 6),
                    "Không tạo đóng góp khi quỹ đã hết hạn (nếu có hạn)",
                    "1. Đăng nhập.\n2. Chọn quỹ đã quá hạn.\n3. Thử đóng góp.",
                    "Bị chặn thao tác.\nThông báo quỹ đã hết hạn.\nKhông tạo contribution.",
                    "Có quỹ quá hạn.",
                ),
            ],
        )
    )

    scenarios.append(
        (
            "Scenario 3- PayOS return / webhook",
            [
                TC(
                    tid(3, 1),
                    "Webhook PayOS: body rỗng -> BadRequest",
                    "1. Gửi request webhook với body rỗng.\n2. Quan sát response.",
                    "API trả 400 BadRequest.\nKhông cập nhật contribution/fund.",
                    "Có endpoint webhook.",
                ),
                TC(
                    tid(3, 2),
                    "Webhook PayOS: signature sai -> BadRequest",
                    "1. Gửi webhook với signature sai.\n2. Quan sát response.",
                    "API trả 400 BadRequest.\nKhông cập nhật dữ liệu.",
                    "Có secret webhook đúng.",
                ),
                TC(
                    tid(3, 3),
                    "Webhook PayOS: code != 00 -> không finalize",
                    "1. Gửi webhook với code khác '00'.\n2. Quan sát transaction.",
                    "API trả 200/OK (hoặc theo spec) nhưng transaction không được đánh dấu paid.\nKhông cộng tiền vào quỹ.",
                    "Có contribution PENDING.",
                ),
                TC(
                    tid(3, 4),
                    "Webhook PayOS: success -> contribution PAID và quỹ tăng số dư",
                    "1. Tạo contribution PENDING.\n2. Gửi webhook success cho đúng orderCode.\n3. Reload quỹ và history.",
                    "Contribution chuyển sang PAID/SUCCESS.\nQuỹ tăng CurrentBalance đúng số tiền.\nHistory có bản ghi mới.",
                    "Có contribution PENDING + webhook hợp lệ.",
                ),
                TC(
                    tid(3, 5),
                    "PayOS Return: truy cập return khi orderCode không tồn tại",
                    "1. Gọi return URL với orderCode không tồn tại.\n2. Quan sát response.",
                    "Trả NotFound hoặc thông báo giao dịch không tồn tại.",
                    "Có return endpoint.",
                ),
            ],
        )
    )

    scenarios.append(
        (
            "Scenario 4- Approve/Reject Fund",
            [
                TC(
                    tid(4, 1),
                    "Top manager approve quỹ PENDING thành công",
                    "1. Đăng nhập user top manager.\n2. Mở danh sách quỹ PENDING.\n3. Approve.",
                    "Quỹ chuyển status = APPROVED.\nHiển thị thông báo thành công.",
                    "Có quỹ PENDING.\nUser role top manager.",
                ),
                TC(
                    tid(4, 2),
                    "Top manager reject quỹ PENDING thành công",
                    "1. Đăng nhập top manager.\n2. Chọn quỹ PENDING.\n3. Reject.",
                    "Quỹ chuyển status = REJECTED (hoặc trạng thái tương ứng).\nHiển thị thông báo.",
                    "Có quỹ PENDING.",
                ),
                TC(
                    tid(4, 3),
                    "Không duyệt được khi không phải top manager",
                    "1. Đăng nhập vice manager/member.\n2. Thử approve quỹ PENDING.",
                    "Bị chặn quyền (403).\nKhông đổi trạng thái quỹ.",
                    "Có quỹ PENDING.",
                ),
                TC(
                    tid(4, 4),
                    "Không duyệt được khi quỹ không tồn tại",
                    "1. Đăng nhập top manager.\n2. Approve với fundId không tồn tại.",
                    "Trả NotFound.\nKhông có thay đổi.",
                    "User top manager.",
                ),
            ],
        )
    )

    scenarios.append(
        (
            "Scenario 5- View/History/Capabilities",
            [
                TC(
                    tid(5, 1),
                    "Xem chi tiết quỹ: fundId tồn tại -> hiển thị đúng",
                    "1. Đăng nhập member của CLB.\n2. Mở quỹ theo fundId.\n3. Quan sát thông tin.",
                    "Hiển thị đúng tên quỹ, số dư, status, ngày hết hạn (nếu có).",
                    "User có quyền xem quỹ.",
                ),
                TC(
                    tid(5, 2),
                    "Xem chi tiết quỹ: fundId không tồn tại -> NotFound",
                    "1. Đăng nhập.\n2. Mở fundId không tồn tại.",
                    "Trả NotFound hoặc thông báo quỹ không tồn tại.",
                    "User đăng nhập.",
                ),
                TC(
                    tid(5, 3),
                    "History paging: page/pageSize hợp lệ -> trả đúng số lượng",
                    "1. Đăng nhập.\n2. Mở lịch sử quỹ.\n3. Set page=1,pageSize=10.\n4. Quan sát kết quả.",
                    "Danh sách trả về đúng số lượng.\nCó tổng số trang/tổng bản ghi nếu API hỗ trợ.",
                    "Quỹ có lịch sử giao dịch.",
                ),
                TC(
                    tid(5, 4),
                    "Capabilities: member thường không có quyền approve/create",
                    "1. Đăng nhập member thường.\n2. Mở capabilities.\n3. Quan sát flags.",
                    "Flags thể hiện không được tạo/duyệt quỹ.\nChỉ được xem/đóng góp theo rule.",
                    "User là member thường.",
                ),
            ],
        )
    )

    return scenarios


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    path = repo_root / "System Test.xlsx"

    import sys

    if len(sys.argv) >= 2:
        path = Path(sys.argv[1])

    wb = load_workbook(path, data_only=False)

    # Remove existing sheet if present (re-generate).
    if "Club Fund" in wb.sheetnames:
        wb.remove(wb["Club Fund"])

    template = wb["Template"]
    ws = wb.copy_worksheet(template)
    ws.title = "Club Fund"

    # Header: workflow name and requirement.
    ws["B1"].value = "Club Fund"
    ws["B2"].value = (
        "System test luồng Club Fund: tạo quỹ, đóng góp, PayOS payment/webhook/return, duyệt quỹ, "
        "xem lịch sử và quyền (capabilities)."
    )

    # Update formulas that count MR_ to CF_
    ws["B3"].value = '=COUNTIF(A11:A1000,"*CF_*")'

    # Clear existing sample rows starting from row 10 down to ~200 in cols A..F (keep round columns pending).
    for r in range(10, 220):
        for c in range(1, 7):  # A..F
            ws.cell(r, c).value = None

    scenarios = build_cases()
    row = 10
    for title, cases in scenarios:
        ws.cell(row, 1).value = title
        row += 1
        for tc in cases:
            ws.cell(row, 1).value = tc.tc_id
            ws.cell(row, 2).value = tc.desc
            ws.cell(row, 3).value = tc.steps
            ws.cell(row, 4).value = tc.expected
            ws.cell(row, 5).value = tc.pre
            # Evidence left blank
            ws.cell(row, 6).value = None
            # Rounds default Pending (keep existing default style if any; set value).
            ws.cell(row, 7).value = "Pending"
            ws.cell(row, 10).value = "Pending"
            ws.cell(row, 13).value = "Pending"
            row += 1

    # Update Test Cases list sheet.
    tcs = wb["Test Cases"]
    insert_row = 9
    while insert_row <= tcs.max_row and tcs.cell(insert_row, 1).value not in (None, ""):
        insert_row += 1
    tcs.cell(insert_row, 1).value = "6.0"
    tcs.cell(insert_row, 2).value = "Club Fund"
    tcs.cell(insert_row, 3).value = "Club Fund"
    tcs.cell(insert_row, 4).value = "Workflow system test cho Club Fund"
    tcs.cell(insert_row, 5).value = "Có dữ liệu CLB, role, quỹ và PayOS sandbox"

    wb.save(path)
    print(f"updated {path} with sheet 'Club Fund' ({sum(len(x[1]) for x in scenarios)} TCs)")


if __name__ == "__main__":
    main()

