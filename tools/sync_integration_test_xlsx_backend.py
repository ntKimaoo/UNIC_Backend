# -*- coding: utf-8 -*-
"""Align Integration Test.xlsx Fund Management sheet with UNIC_Backend ClubFundController + ClubFundService."""
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
XLSX = REPO / "Integration Test.xlsx"

# Row with Test requirement is B2 on sheet Fund Management
UPDATES_BY_TC_ID: dict[str, dict[str, str]] = {
    "FM_0101": {
        "B": "Verify successful fund creation with createfinance and Manager or Vice club role",
        "C": (
            "1. Log in as club member with createfinance policy and ClubRole level 1 or 2.\n"
            "2. POST /api/clubs/{clubId}/funds with body FundName, optional Description, optional ExpiresAt. No initial balance field.\n"
            "3. Validate response and DB."
        ),
        "D": (
            "HTTP 200, success true, data contains created fund.\n"
            "If creator is top manager ClubRole level 1: new fund status APPROVED, ApprovedBy set to creator.\n"
            "If creator is Vice Manager level 2: new fund status PENDING, ApprovedBy null until approval."
        ),
        "E": "User is ACTIVE club member with createfinance. ClubRole is Manager level 1 or Vice level 2 per ClubFundService.",
    },
    "FM_0102": {
        "B": "Verify fund creation is denied without permission or wrong role",
        "C": (
            "1. Case A: member without createfinance — POST /api/clubs/{clubId}/funds.\n"
            "2. Case B: member with createfinance but club role below Manager or Vice — POST same.\n"
            "3. Observe HTTP result."
        ),
        "D": (
            "HTTP 403, success false.\n"
            "Case A: policy handler fails before service.\n"
            "Case B: UnauthorizedAccessException from service — only Manager or Vice may create funds.\n"
            "No new fund in DB."
        ),
        "E": "Target club exists. User is member; Case A lacks createfinance. Case B has createfinance but role is not level 1 or 2.",
    },
    "FM_0103": {
        "B": "Verify fund approval by top club manager with editfinance",
        "C": (
            "1. Fund in PENDING exists.\n"
            "2. Log in as user with editfinance AND ClubRole level 1 only top manager per ApproveFundAsync.\n"
            "3. POST /api/clubs/{clubId}/funds/approve with fundId and action APPROVE."
        ),
        "D": (
            "HTTP 200, success true.\n"
            "Fund status APPROVED in DB. ApprovedBy set to approver."
        ),
        "E": "PENDING fund. User has editfinance and ClubRole level 1 in that club. ApproveFundAsync does not bypass club role for JWT Admin alone; user must be member with top manager level.",
    },
    "FM_0104": {
        "B": "Verify approve or reject fails for invalid fund state or invalid reject payload",
        "C": (
            "1. Fund already APPROVED or REJECTED — POST approve with action APPROVE or REJECT as top manager with editfinance — expect 400.\n"
            "2. Fund PENDING — POST reject with action REJECT but omit rejectReason or reason under 5 characters — expect 400 ArgumentException from service.\n"
            "3. Fund PENDING — POST reject with action REJECT and valid rejectReason length 5 to 2000 — expect 200 and REJECTED state."
        ),
        "D": (
            "Step 1: HTTP 400, message explains invalid state. No DB change.\n"
            "Step 2: HTTP 400, Vietnamese validation message for missing or short rejectReason per ClubFundService.\n"
            "Step 3: HTTP 200, fund REJECTED with RejectReason and RejectedAt UTC stored."
        ),
        "E": "Funds in APPROVED, REJECTED, and PENDING. Approver is top manager with editfinance for each POST.",
    },
    "FM_0201": {
        "B": "Verify get fund list by club with server-side status filtering",
        "C": (
            "1. Log in with viewfinance.\n"
            "2. GET /api/clubs/{clubId}/funds with optional query status, page, pageSize, search, sort.\n"
            "3. Repeat with user who is not system Admin and not top manager with editfinance — try status PENDING or ALL."
        ),
        "D": (
            "HTTP 200, paged data, club scoped.\n"
            "Users without Admin and without top manager plus editfinance receive only APPROVED funds; status query is overridden server-side.\n"
            "Privileged users may filter ALL, PENDING, APPROVED, REJECTED per GetFundsByClubIdPagedAsync."
        ),
        "E": "Mixed fund statuses exist. Test both privileged and ordinary member with viewfinance.",
    },
    "FM_0202": {
        "C": (
            "1. Log in with viewfinance.\n"
            "2. GET /api/clubs/{clubId}/funds?page=0&pageSize=9.\n"
            "3. GET /api/clubs/{clubId}/funds?page=1&pageSize=101."
        ),
        "D": "HTTP 400 for invalid page or pageSize. Messages per controller: page at least 1, pageSize 1 to 100.",
    },
    "FM_0203": {
        "C": (
            "1. GET /api/clubs/{clubId}/funds/{fundId} for existing fund in that club.\n"
            "2. GET same route with non-existent fundId.\n"
            "3. Optional: existing fundId but user not allowed access to club returns 403 per CanAccessClubAsync."
        ),
        "D": (
            "Existing fund: HTTP 200, success true, data includes rejectReason and rejectionReasonVi when REJECTED.\n"
            "Non-existent fund: HTTP 404, success false."
        ),
    },
    "FM_0301": {
        "C": (
            "1. Log in as ACTIVE club member.\n"
            "2. POST /api/clubs/{clubId}/funds/contribute with FundId and Amount at least 1000 VND.\n"
            "3. Validate ContributeResponseDto fields."
        ),
        "D": (
            "HTTP 200, success true.\n"
            "Response data includes transactionId, checkoutUrl, qrCode, paymentLinkId, amount, paymentLinkExpiresAtUtc per API serialization."
        ),
        "E": "Fund status APPROVED, not expired. User member of fund club.",
    },
    "FM_0302": {
        "D": "HTTP 400 from model validation or service. Minimum amount 1,000 VND. No transaction persisted.",
    },
    "FM_0401": {
        "C": (
            "1. Create member contribution so transaction exists.\n"
            "2. GET /api/clubs/{clubId}/funds/contribute/{transactionId}/status as owning user.\n"
            "3. Call with wrong club or other user transaction — expect 404 or null handling per service."
        ),
    },
    "FM_0402": {
        "C": (
            "1. Log in with viewfinance.\n"
            "2. GET /api/clubs/{clubId}/funds/history/{fundId}?status=APPROVED&scope=mine&page=1&pageSize=10.\n"
            "3. Validate paging and filters match GetFundHistoryPagedAsync."
        ),
    },
    "FM_0403": {
        "C": (
            "1. GET /api/clubs/{clubId}/funds/report-summary?fromUtc=...&toUtc=...\n"
            "2. GET /api/clubs/{clubId}/funds/transactions?fromUtc=...&toUtc=...&page=1&pageSize=20.\n"
            "3. Validate totals and list. Note transactions pageSize max 100 per controller."
        ),
        "D": "HTTP 200 when date range valid and user has club access and viewfinance. Invalid date range returns structured 400 per BuildBadRequest.",
    },
}


def _patch_test_requirement(ws) -> None:
    cell = ws["B2"]
    v = cell.value
    if not isinstance(v, str):
        return
    tag = "Backend note:"
    if tag in v:
        return
    cell.value = (
        v.rstrip()
        + " "
        + tag
        + " Approve or reject fund entity requires ClubRole level 1 plus editfinance. Reject requires rejectReason 5 to 2000 characters. "
        "GET fund list overrides status filter to APPROVED-only for users without Admin and without that approve privilege. "
        "GET /api/clubs/{clubId}/funds/my defaults mineType CREATED for creator-scoped list."
    )


def main() -> None:
    if not XLSX.is_file():
        raise SystemExit(f"Missing {XLSX}")

    wb = openpyxl.load_workbook(XLSX)
    if "Fund Management" not in wb.sheetnames:
        raise SystemExit("Sheet 'Fund Management' not found")
    ws = wb["Fund Management"]
    _patch_test_requirement(ws)

    for row in ws.iter_rows(min_row=10, max_row=500):
        tc = row[0].value
        if not tc or not str(tc).startswith("FM_"):
            continue
        key = str(tc).strip()
        if key not in UPDATES_BY_TC_ID:
            continue
        patch = UPDATES_BY_TC_ID[key]
        # A=0, B=1, C=2, D=3, E=4
        for letter, idx in (("B", 1), ("C", 2), ("D", 3), ("E", 4)):
            if letter in patch:
                row[idx].value = patch[letter]

    wb.save(XLSX)
    print(f"Updated: {XLSX}")


if __name__ == "__main__":
    main()
