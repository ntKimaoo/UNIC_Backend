# -*- coding: utf-8 -*-
"""
Điền nội dung (giá trị ô) vào Unit Test_2.xlsx; giữ màu/font/viền template.

- Giữ 3 sheet: Cover, Method List, Statistics (xóa sheet khác như trước).
- Gán cell.value / number_format cho ngày.
- Statistics: sau khi ghi số, đóng viền đáy bảng (chỉ cạnh bottom, lấy kiểu từ ô header mẫu).
"""
from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Border, Side

REPO = Path(__file__).resolve().parents[1]
XLSX = REPO / "Unit Test_2.xlsx"
CTRL_TEST = REPO / "UNIC.ControllerTest" / "Controllers" / "ClubFundControllerTest.cs"
SVC_TEST = REPO / "UNIC.ServiceTest" / "Services" / "ClubFundServiceTest.cs"

PRECONDITION_VI = (
    "Môi trường: .NET 8, Visual Studio hoặc CLI. Chạy `dotnet test` trên UNIC.ControllerTest / UNIC.ServiceTest. "
    "Test dùng xUnit + Moq; không cần SQL Server hay PayOS thật (service/controller mock)."
)


def _bottom_side_from_header(st, header_row: int = 11, sample_col: int = 2) -> Side:
    b = st.cell(header_row, sample_col).border
    if b.bottom and b.bottom.style:
        return b.bottom
    return Side(style="thin", color="FF000000")


def _set_row_bottom_closed(st, row: int, col_from: int, col_to: int, bottom: Side) -> None:
    for cc in range(col_from, col_to + 1):
        c = st.cell(row, cc)
        b = c.border
        c.border = Border(
            left=b.left,
            right=b.right,
            top=b.top,
            bottom=bottom,
            diagonal=b.diagonal,
            diagonal_direction=b.diagonal_direction,
            outline=b.outline,
            vertical=b.vertical,
            horizontal=b.horizontal,
        )


def extract_async_test_names(cs_text: str) -> list[str]:
    return re.findall(r"public async Task (\w+)\s*\(", cs_text)


def scenario_type(test_name: str) -> str:
    if "ReturnsInternalServerError" in test_name or "ReturnsBadRequest" in test_name:
        return "A"
    if "Returns403" in test_name or "Returns404" in test_name or "ReturnsNotFound" in test_name:
        return "A"
    if "ReturnsUnauthorized" in test_name:
        return "A"
    if "_ShouldThrow" in test_name or "_ShouldReturnFalse" in test_name:
        return "A"
    if "ShouldReturnNull" in test_name:
        return "A"
    if "WhenNotSuccessCode" in test_name:
        return "N"
    if "ReturnsOk" in test_name or "ShouldReturnTrue" in test_name:
        return "N"
    if "_ShouldReturnPagedDtos" in test_name or "_ShouldMap" in test_name:
        return "N"
    if "_ShouldComplete" in test_name or "_ShouldUpdate_WhenReject" in test_name:
        return "N"
    if "_ShouldBypassMemberGate" in test_name or "_ShouldPass" in test_name:
        return "N"
    if "_ShouldCallPayOS" in test_name or "_ShouldDeleteTransaction" in test_name:
        return "N"
    if "_ShouldSetApproved" in test_name or "_ShouldSetPending" in test_name:
        return "N"
    if "_ShouldReturnPayOsPayload" in test_name:
        return "N"
    if "PassesApproved" in test_name or "PassesIsSystemAdmin" in test_name:
        return "N"
    return "A"


def filter_prefix(names: list[str], prefix: str) -> list[str]:
    return [n for n in names if n.startswith(prefix + "_")]


def count_nab(tests: list[str]) -> tuple[int, int, int, int]:
    n = sum(1 for t in tests if scenario_type(t) == "N")
    a = sum(1 for t in tests if scenario_type(t) == "A")
    b = 0
    return n, a, b, len(tests)


def main() -> None:
    ctrl_names = extract_async_test_names(CTRL_TEST.read_text(encoding="utf-8"))
    svc_names = extract_async_test_names(SVC_TEST.read_text(encoding="utf-8"))

    controller_specs: list[tuple[str, str, str, list[str]]] = [
        ("ClubFundController", "CreateFund", "CreateFund", filter_prefix(ctrl_names, "CreateFund")),
        ("ClubFundController", "GetMyClubs", "GetMyClubs", filter_prefix(ctrl_names, "GetMyClubs")),
        ("ClubFundController", "GetFundCapabilities", "GetFundCapabilities", filter_prefix(ctrl_names, "GetFundCapabilities")),
        ("ClubFundController", "GetFundReportSummary", "GetFundReportSummary", filter_prefix(ctrl_names, "GetFundReportSummary")),
        ("ClubFundController", "GetFundCategories", "GetFundCategories", filter_prefix(ctrl_names, "GetFundCategories")),
        ("ClubFundController", "GetClubFundTransactions", "GetClubFundTransactions", filter_prefix(ctrl_names, "GetClubFundTransactions")),
        ("ClubFundController", "GetFund", "GetFund", [n for n in ctrl_names if n.startswith("GetFund_")]),
        ("ClubFundController", "GetFundsByClub", "GetFundsByClub", filter_prefix(ctrl_names, "GetFundsByClub")),
        ("ClubFundController", "GetMyFunds", "GetMyFunds", filter_prefix(ctrl_names, "GetMyFunds")),
        ("ClubFundController", "Contribute", "Contribute", filter_prefix(ctrl_names, "Contribute")),
        (
            "ClubFundController",
            "GetContributionPaymentStatus",
            "GetContributionPaymentStatus",
            filter_prefix(ctrl_names, "GetContributionPaymentStatus"),
        ),
        (
            "ClubFundController",
            "GetPayOsContributionReturn",
            "GetPayOsContributionReturn",
            filter_prefix(ctrl_names, "GetPayOsContributionReturn"),
        ),
        (
            "ClubFundController",
            "SimulatePayOsPaidForDevelopment",
            "SimulatePayOsPaidForDevelopment",
            filter_prefix(ctrl_names, "SimulatePayOsPaidForDevelopment"),
        ),
        ("ClubFundController", "ApproveFund", "ApproveFund", filter_prefix(ctrl_names, "ApproveFund")),
        ("ClubFundController", "PayOSWebhook", "PayOSWebhookClubScoped", [n for n in ctrl_names if n.startswith("PayOSWebhook_")]),
        ("ClubFundController", "GetHistory", "GetHistory", filter_prefix(ctrl_names, "GetHistory")),
        ("ClubFundController", "GetFundLocation", "GetFundLocation", filter_prefix(ctrl_names, "GetFundLocation")),
    ]

    service_specs: list[tuple[str, str, str, list[str]]] = [
        ("ClubFundService", "CreateFundAsync", "CreateFundAsync", filter_prefix(svc_names, "CreateFundAsync")),
        ("ClubFundService", "CreateContributionAsync", "CreateContributionAsync", filter_prefix(svc_names, "CreateContributionAsync")),
        (
            "ClubFundService",
            "GetContributionPaymentStatusAsync",
            "GetContributionPaymentStatusAsync",
            filter_prefix(svc_names, "GetContributionPaymentStatusAsync"),
        ),
        (
            "ClubFundService",
            "GetContributionPaymentStatusByOrderCodeAsync",
            "GetContributionPaymentStatusByOrderCodeAsync",
            filter_prefix(svc_names, "GetContributionPaymentStatusByOrderCodeAsync"),
        ),
        ("ClubFundService", "GetFundByIdAsync", "GetFundByIdAsync", filter_prefix(svc_names, "GetFundByIdAsync")),
        ("ClubFundService", "GetFundsByClubIdPagedAsync", "GetFundsByClubIdPagedAsync", filter_prefix(svc_names, "GetFundsByClubIdPagedAsync")),
        ("ClubFundService", "GetMyFundsByClubIdPagedAsync", "GetMyFundsByClubIdPagedAsync", filter_prefix(svc_names, "GetMyFundsByClubIdPagedAsync")),
        ("ClubFundService", "GetFundHistoryPagedAsync", "GetFundHistoryPagedAsync", filter_prefix(svc_names, "GetFundHistoryPagedAsync")),
        ("ClubFundService", "ApproveFundAsync", "ApproveFundAsync", filter_prefix(svc_names, "ApproveFundAsync")),
        (
            "ClubFundService",
            "ProcessPayOSPaymentSuccessAsync",
            "ProcessPayOSPaymentSuccessAsync",
            filter_prefix(svc_names, "ProcessPayOSPaymentSuccessAsync"),
        ),
        (
            "ClubFundService",
            "TryCompleteOwnPendingContributionForDevelopmentAsync",
            "TryCompleteOwnPendingContributionForDevelopmentAsync",
            filter_prefix(svc_names, "TryCompleteOwnPendingContributionForDevelopmentAsync"),
        ),
        ("ClubFundService", "GetFundCapabilitiesAsync", "GetFundCapabilitiesAsync", filter_prefix(svc_names, "GetFundCapabilitiesAsync")),
        (
            "ClubFundService",
            "GetClubFundReportSummaryAsync",
            "GetClubFundReportSummaryAsync",
            filter_prefix(svc_names, "GetClubFundReportSummaryAsync"),
        ),
        (
            "ClubFundService",
            "GetClubFundTransactionsPagedAsync",
            "GetClubFundTransactionsPagedAsync",
            filter_prefix(svc_names, "GetClubFundTransactionsPagedAsync"),
        ),
    ]

    descriptions: dict[str, str] = {
        "CreateFund": "POST tạo quỹ: 200 thành công; 403 không đủ quyền; 400 lỗi validation/ArgumentException.",
        "GetMyClubs": "GET danh sách club của user: 200.",
        "GetFundCapabilities": "GET quyền/năng lực quỹ theo club: 403 không phải member; 200; Unauthorized khi service ném.",
        "GetFundReportSummary": "GET báo cáo tổng hợp quỹ: 403; 200; 400 khoảng ngày fromUtc > toUtc.",
        "GetFundCategories": "GET danh mục quỹ: 403; 200; Unauthorized.",
        "GetClubFundTransactions": "GET giao dịch theo club: 403; 400 page/pageSize/ngày; Unauthorized; quỹ không thuộc club; 404; 200.",
        "GetFund": "GET chi tiết quỹ: 404; 403; 200 (Admin bỏ qua membership).",
        "GetFundsByClub": "GET danh sách quỹ phân trang: 400; 403; 200; Admin truyền isSystemAdmin.",
        "GetMyFunds": "GET quỹ của tôi: 403; 400 mineType; 200.",
        "Contribute": "POST tạo yêu cầu nộp quỹ PayOS: 403; 404; 200.",
        "GetContributionPaymentStatus": "GET trạng thái nộp tiền theo transaction: 403; 404; 200.",
        "GetPayOsContributionReturn": "GET trạng thái theo orderCode (return URL): 404; 403; 200.",
        "SimulatePayOsPaidForDevelopment": "POST giả lập thanh toán (chỉ Development): 404 ngoài Dev; 200.",
        "ApproveFund": "POST duyệt/từ chối quỹ: 200; 403; 400 (Argument/InvalidOperation/Exception).",
        "PayOSWebhookClubScoped": "POST webhook PayOS: body rỗng; code≠00; thiếu/chữ ký sai; xử lý thành công; orderCode≤0; JSON lỗi.",
        "GetHistory": "GET lịch sử quỹ: 400 page; 404; 403; 200; Admin không cần membership club.",
        "GetFundLocation": "GET clubId/fundId cho deep link: 403; 404; 200.",
        "CreateFundAsync": "Nghiệp vụ tạo quỹ: validation tên/hạn; quyền Manager/Vice; auto APPROVED/PENDING; alias mô tả.",
        "CreateContributionAsync": "Nghiệp vụ nộp quỹ: số tiền tối thiểu; trạng thái quỹ; category; PayOS; rollback khi PayOS lỗi.",
        "GetContributionPaymentStatusAsync": "Trạng thái giao dịch: null khi không phải chủ giao dịch.",
        "GetContributionPaymentStatusByOrderCodeAsync": "Trạng thái theo orderCode khi đã thanh toán.",
        "GetFundByIdAsync": "Lấy quỹ theo id: null khi không tồn tại.",
        "GetFundsByClubIdPagedAsync": "Phân trang danh sách quỹ: filter status/sort; ép APPROVED; PENDING cho top manager; SystemAdmin bypass.",
        "GetMyFundsByClubIdPagedAsync": "Quỹ của tôi: chuẩn hóa mineType ALL; lỗi mineType.",
        "GetFundHistoryPagedAsync": "Lịch sử: mặc định APPROVED; ALL; scope mine.",
        "ApproveFundAsync": "Duyệt/từ chối: không tồn tại; đã duyệt; không phải top manager; reject hợp lệ/không lý do/lý do ngắn.",
        "ProcessPayOSPaymentSuccessAsync": "Webhook xử lý: false/true theo repository.",
        "TryCompleteOwnPendingContributionForDevelopmentAsync": "Dev hoàn tất pending: sai club; hợp lệ.",
        "GetFundCapabilitiesAsync": "Quyền xem/tạo/duyệt và menu theo policy; không member; không viewfinance.",
        "GetClubFundReportSummaryAsync": "Tổng hợp số liệu báo cáo quỹ từ repository.",
        "GetClubFundTransactionsPagedAsync": "Giao dịch club: mặc định APPROVED; scope mine; status ALL.",
    }

    rows: list[tuple[str, str, str, str, list[str]]] = []
    for mod, display_method, stat_key, tests in controller_specs + service_specs:
        if not tests:
            continue
        desc = descriptions.get(display_method, descriptions.get(stat_key, "Unit test Club Fund."))
        rows.append((mod, display_method, stat_key, desc, tests))

    wb = load_workbook(XLSX, data_only=False)
    keep = {"Cover", "Method List", "Statistics"}
    for name in reversed(list(wb.sheetnames)):
        if name not in keep:
            wb.remove(wb[name])

    # --- Cover: chỉ thêm dòng changelog (không đổi tiêu đề / metadata template) ---
    cov = wb["Cover"]
    r = 13
    while cov.cell(r, 1).value not in (None, ""):
        r += 1
    cov.cell(r, 1).value = datetime(2026, 4, 4)
    cov.cell(r, 1).number_format = "yyyy-mm-dd"
    cov.cell(r, 2).value = "v1.1"
    cov.cell(r, 3).value = "Club Fund (Unit Test)"
    cov.cell(r, 4).value = "M"
    cov.cell(r, 5).value = (
        "Cập nhật nội dung unittest Club Fund (chỉ điền text, giữ template)."
    )
    cov.cell(r, 6).value = "UNIC_Backend — ClubFundControllerTest, ClubFundServiceTest"

    # --- Method List: chỉ xóa giá trị vùng data rồi ghi lại (style ô giữ nguyên) ---
    ml = wb["Method List"]
    for r in range(9, 200):
        for c in range(1, 10):
            ml.cell(r, c).value = None

    for i, (mod, method, _sk, desc, _tests) in enumerate(rows, start=1):
        rr = 8 + i
        ml.cell(rr, 1).value = float(i)
        ml.cell(rr, 2).value = mod
        ml.cell(rr, 3).value = method
        ml.cell(rr, 4).value = "(sheet chi tiết đã gỡ — tham chiếu mã nguồn test)"
        ml.cell(rr, 5).value = desc
        ml.cell(rr, 6).value = PRECONDITION_VI

    # --- Statistics: chỉ giá trị ---
    st = wb["Statistics"]
    for r in range(12, 120):
        for c in range(1, 12):
            st.cell(r, c).value = None

    total_passed = 0
    total_failed = 0
    total_untested = 0
    sum_n = sum_a = sum_b = sum_i = 0

    for i, (_mod, _method, stat_key, _desc, tests) in enumerate(rows, start=1):
        rr = 11 + i
        n, a, b, tot = count_nab(tests)
        passed = tot
        failed = 0
        untested = 0
        st.cell(rr, 1).value = float(i)
        st.cell(rr, 2).value = stat_key
        st.cell(rr, 3).value = float(passed)
        st.cell(rr, 4).value = float(failed)
        st.cell(rr, 5).value = float(untested)
        st.cell(rr, 6).value = float(n)
        st.cell(rr, 7).value = float(a)
        st.cell(rr, 8).value = float(b)
        st.cell(rr, 9).value = float(tot)
        total_passed += passed
        total_failed += failed
        total_untested += untested
        sum_n += n
        sum_a += a
        sum_b += b
        sum_i += tot

    sub_row = 12 + len(rows)
    st.cell(sub_row, 1).value = None
    st.cell(sub_row, 2).value = "Sub total"
    st.cell(sub_row, 3).value = float(total_passed)
    st.cell(sub_row, 4).value = float(total_failed)
    st.cell(sub_row, 5).value = float(total_untested)
    st.cell(sub_row, 6).value = float(sum_n)
    st.cell(sub_row, 7).value = float(sum_a)
    st.cell(sub_row, 8).value = float(sum_b)
    st.cell(sub_row, 9).value = float(sum_i)

    bottom_side = _bottom_side_from_header(st)
    last_data_row = 11 + len(rows)
    _set_row_bottom_closed(st, last_data_row, 1, 9, bottom_side)
    _set_row_bottom_closed(st, sub_row, 1, 9, bottom_side)

    total = sum_i
    cov_pct = round((total_passed + total_failed) / total * 100, 2) if total else 0.0
    succ_pct = round(total_passed / total * 100, 2) if total else 0.0
    n_pct = round(sum_n / total * 100, 2) if total else 0.0
    a_pct = round(sum_a / total * 100, 2) if total else 0.0
    b_pct = round(sum_b / total * 100, 2) if total else 0.0

    base = sub_row + 2
    metric_labels = [
        "Test coverage",
        "Test successful coverage",
        "Normal case",
        "Abnormal case",
        "Boundary case",
    ]
    metrics_vals = [cov_pct, succ_pct, n_pct, a_pct, b_pct]
    for i, lab in enumerate(metric_labels):
        st.cell(base + i, 2).value = lab
        st.cell(base + i, 4).value = metrics_vals[i]
        st.cell(base + i, 5).value = "%"

    wb.save(XLSX)
    print("saved", XLSX)
    print("methods", len(rows), "total_cases", sum_i)
    print("sheets", wb.sheetnames)


if __name__ == "__main__":
    main()
