from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def main() -> None:
    path = Path(__file__).resolve().parents[1] / "Unit Test.xlsx"
    wb = load_workbook(path, data_only=False)
    print(f"sheets: {len(wb.sheetnames)}")

    club_sheets = [
        s
        for s in wb.sheetnames
        if ("ClubFund" in s) or ("Club.Fund" in s) or ("Club Fund" in s)
    ]
    print("club sheets:")
    for s in club_sheets:
        print(f"- {s}")

    for name in club_sheets:
        ws = wb[name]
        max_row, max_col = ws.max_row, ws.max_column
        weird: list[tuple[str, str]] = []

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str) and len(v) > 30:
                    weird.append((ws.cell(r, c).coordinate, v))

        print(f"\n== {name} used_range=({max_row}x{max_col}) long_text_cells={len(weird)}")
        for addr, v in weird[:80]:
            preview = v.replace("\r\n", "\\n").replace("\n", "\\n")
            if len(preview) > 120:
                preview = preview[:120] + "…"
            print(f"{addr}: {preview}")


if __name__ == "__main__":
    main()

