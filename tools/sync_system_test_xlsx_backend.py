# -*- coding: utf-8 -*-
"""Align System Test.xlsx Club Fund sheet with current UNIC_Backend Club Fund API."""
from __future__ import annotations

from copy import copy
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[1]
XLSX = REPO / "System Test.xlsx"

# Verbatim backend messages — keep (rejectReason) inside Vietnamese string as in ClubFundService.
_MSG_REJECT_REQUIRED = "Khi từ chối quỹ, bắt buộc nhập lý do (rejectReason)."
_MSG_REJECT_MIN5 = "Lý do từ chối phải có ít nhất 5 ký tự sau khi bỏ khoảng trắng đầu cuối."
_MSG_REJECT_MAX2000 = "Lý do từ chối không được vượt quá 2000 ký tự."

# column A = Test Case ID -> columns B–E
UPDATES: dict[str, dict[str, str]] = {
    "CF_3003_1.01": {
        "B": "Fund created successfully when user is top-level Manager status APPROVED",
        "C": (
            "1. Log in as club Manager level 1.\n"
            "2. Open create fund POST .../funds with createfinance policy.\n"
            "3. Enter valid FundName, optional Description, optional ExpiresAt as contribution deadline only. "
            "API does not accept an initial balance field.\n"
            "4. Submit."
        ),
        "D": (
            "Fund created successfully.\n"
            "Fund is created with status APPROVED. CreatedBy and ApprovedBy set to creator.\n"
            "Success response per API. FE may redirect to fund details or list."
        ),
        "E": (
            "User belongs to the club with Manager role level 1, highest level in club.\n"
            "Club exists. Membership ACTIVE. User has createfinance policy."
        ),
    },
    "CF_3003_1.02": {
        "E": (
            "User belongs to the club with Vice Manager role level 2, ACTIVE.\n"
            "User has createfinance policy together with Manager or Vice role to create fund."
        ),
    },
    "CF_3003_1.04": {
        "B": "Cannot create fund when fund name duplicates an existing non-rejected fund in the same club",
        "C": (
            "1. Log in as authorized user, Manager or Vice with createfinance.\n"
            "2. Create a fund with FundName equal to an existing fund in the club that is not REJECTED.\n"
            "3. Submit."
        ),
        "D": "Validation error duplicate fund name in club.\nFund is not created.",
        "E": "Another fund with the same normalized name exists in the club, PENDING or APPROVED.",
    },
    "CF_3003_2.01": {
        "D": (
            "Contribution record created. Status PENDING until PayOS payment completes.\n"
            "Response includes CheckoutUrl, QrCode, PaymentLinkId, TransactionId as PayOS order code, Amount, PaymentLinkExpiresAtUtc."
        ),
        "E": (
            "Fund exists. Status APPROVED. Not past ExpiresAt if set.\n"
            "Member ACTIVE in club. PayOS sandbox configured.\n"
            "Minimum contribution amount on API is 1,000 VND."
        ),
    },
    "CF_3003_2.03": {
        "E": "Fund exists. API enforces minimum Amount at least 1,000 VND per ContributeRequestDto.",
    },
    "CF_3003_4.02": {
        "C": (
            "1. Log in as top manager level 1 with editfinance policy.\n"
            "2. Select a PENDING fund.\n"
            "3. Reject with action REJECT and rejectReason. Required: non-empty after trim, length 5 to 2000. "
            "JSON may use rejectReason, RejectReason, rejectionReason, or RejectionReason per ApproveFundDto.\n"
            "4. Submit POST .../funds/approve."
        ),
        "D": (
            "Fund status REJECTED.\n"
            "RejectReason, RejectedAt UTC, ApprovedBy stored for audit.\n"
            "Fund API responses include rejectReason, rejectedAt, rejectionReasonVi when status is REJECTED."
        ),
        "E": "A fund in PENDING exists. User is top manager with editfinance.",
    },
    "CF_3003_4.01": {
        "C": (
            "1. Log in as top manager with editfinance.\n"
            "2. Open club fund list GET .../funds. Only system Admin or top manager with editfinance may filter by PENDING. "
            "Other members receive APPROVED-only list enforced server-side.\n"
            "3. Approve with action APPROVE. rejectReason not required."
        ),
        "E": "A fund in PENDING exists. User is top manager with editfinance or system Admin.",
    },
    "CF_3003_4.04": {
        "D": (
            "400 Bad Request with message that the fund does not exist. Approve endpoint returns 400 for missing fund.\n"
            "No fund is updated."
        ),
        "E": "User is top manager with editfinance. fundId does not exist.",
    },
    "CF_3003_4.05": {
        "B": "Cannot reject fund when rejectReason is missing, null, or whitespace-only",
        "C": (
            "1. PENDING fund exists. User is club top manager with editfinance. "
            "Endpoint POST /api/clubs/{clubId}/funds/approve.\n"
            "2. Body includes fundId and action REJECT. Omit rejectReason or send only spaces.\n"
            "3. Observe HTTP status and JSON body."
        ),
        "D": (
            "HTTP 400 Bad Request. ClubFundController maps ArgumentException to BadRequest, success false, message from exception.\n"
            f"ClubFundService message exactly: {_MSG_REJECT_REQUIRED}\n"
            "Fund row unchanged, still PENDING."
        ),
        "E": "Same preconditions as CF_3003_4.02.",
    },
    "CF_3003_4.06": {
        "B": "Cannot reject fund when rejectReason has fewer than 5 characters after trim",
        "C": (
            "1. Same preconditions and endpoint as CF_3003_4.05.\n"
            "2. action REJECT with rejectReason shorter than 5 characters after trim, such as abc.\n"
            "3. Observe response."
        ),
        "D": (
            "HTTP 400 Bad Request. Same controller mapping as CF_3003_4.05.\n"
            f"ClubFundService message exactly: {_MSG_REJECT_MIN5}\n"
            "Fund row unchanged."
        ),
        "E": "Same as CF_3003_4.05.",
    },
    "CF_3003_4.07": {
        "B": "Cannot reject fund when rejectReason exceeds 2000 characters after trim",
        "C": (
            "1. Same preconditions and endpoint as CF_3003_4.05.\n"
            "2. action REJECT with rejectReason longer than 2000 characters after trim per rejectReasonMaxLen in ClubFundService.\n"
            "3. Observe response."
        ),
        "D": (
            "HTTP 400 Bad Request. ClubFundController maps ArgumentException to BadRequest.\n"
            f"ClubFundService message exactly: {_MSG_REJECT_MAX2000}\n"
            "Fund row unchanged."
        ),
        "E": "Same as CF_3003_4.05.",
    },
    "CF_3003_5.01": {
        "D": (
            "Shows fund name, balances, status, ExpiresAt if any.\n"
            "If REJECTED: rejectReason, rejectedAt, rejectionReasonVi with same text as rejectReason for display."
        ),
    },
    "CF_3003_5.04": {
        "D": (
            "Capabilities reflect policies. CanCreateFund requires createfinance and Manager or Vice role level.\n"
            "CanApproveOrRejectFundEntity requires editfinance and top manager level 1 only.\n"
            "CanViewFunds requires viewfinance. CanContribute is true for active members when other rules allow."
        ),
        "E": "Regular member: typically no createfinance or not Vice or Manager. Flags off for create and approve.",
    },
}

PENDING_ROUND_COLS = {7, 10, 13}


def _copy_row_style(ws, src_row: int, dst_row: int, max_col: int = 15) -> None:
    """Match borders, fill, font, alignment to template row so inserted rows match the sheet template."""
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        dst.font = copy(src.font)
        dst.border = copy(src.border)
        dst.fill = copy(src.fill)
        dst.number_format = copy(src.number_format)
        dst.protection = copy(src.protection)
        dst.alignment = copy(src.alignment)


def _apply_reject_rows_style(ws) -> None:
    template_row: int | None = None
    for r in range(10, 400):
        if ws.cell(r, 1).value == "CF_3003_4.04":
            template_row = r
            break
    if template_row is None:
        return
    for tc in ("CF_3003_4.05", "CF_3003_4.06", "CF_3003_4.07"):
        for r in range(10, 400):
            if ws.cell(r, 1).value == tc:
                _copy_row_style(ws, template_row, r)
                break


def _ensure_reject_reason_validation_rows(ws) -> None:
    """Insert CF_3003_4.05 / 4.06 once."""
    for row in ws.iter_rows(min_row=10, max_row=300):
        if row[0].value == "CF_3003_4.05":
            return

    insert_at: int | None = None
    for r in range(10, 300):
        v = ws.cell(r, 1).value
        if v and str(v).startswith("Scenario 5"):
            insert_at = r
            break
    if insert_at is None:
        return

    ws.insert_rows(insert_at, 2)
    u5, u6 = UPDATES["CF_3003_4.05"], UPDATES["CF_3003_4.06"]

    def write_row(r: int, tc_id: str, u: dict[str, str]) -> None:
        ws.cell(row=r, column=1, value=tc_id)
        ws.cell(row=r, column=2, value=u["B"])
        ws.cell(row=r, column=3, value=u["C"])
        ws.cell(row=r, column=4, value=u["D"])
        ws.cell(row=r, column=5, value=u["E"])
        for col in PENDING_ROUND_COLS:
            ws.cell(row=r, column=col, value="Pending")

    write_row(insert_at, "CF_3003_4.05", u5)
    write_row(insert_at + 1, "CF_3003_4.06", u6)


def _ensure_reject_reason_max_length_row(ws) -> None:
    """Insert CF_3003_4.07 after CF_3003_4.06 if missing."""
    for row in ws.iter_rows(min_row=10, max_row=400):
        if row[0].value == "CF_3003_4.07":
            return

    row_406: int | None = None
    for r in range(10, 400):
        if ws.cell(r, 1).value == "CF_3003_4.06":
            row_406 = r
            break
    if row_406 is None:
        return

    insert_at = row_406 + 1
    ws.insert_rows(insert_at, 1)
    u7 = UPDATES["CF_3003_4.07"]
    ws.cell(row=insert_at, column=1, value="CF_3003_4.07")
    ws.cell(row=insert_at, column=2, value=u7["B"])
    ws.cell(row=insert_at, column=3, value=u7["C"])
    ws.cell(row=insert_at, column=4, value=u7["D"])
    ws.cell(row=insert_at, column=5, value=u7["E"])
    for col in PENDING_ROUND_COLS:
        ws.cell(row=insert_at, column=col, value="Pending")


def main() -> None:
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Club Fund"]

    _ensure_reject_reason_validation_rows(ws)
    _ensure_reject_reason_max_length_row(ws)

    req = ws["B2"].value
    marker = "Backend alignment:"
    if isinstance(req, str) and marker not in req:
        ws["B2"].value = (
            req.rstrip(".")
            + ". "
            + marker
            + " GET .../funds enforces APPROVED-only list for users who are not system Admin "
            "and not club top manager with editfinance; status query is overridden. "
            "GET .../funds/my defaults mineType CREATED for funds created by current user. "
            "Reject fund requires rejectReason min 5 chars max 2000. "
            "Create fund has no initial balance field; TotalAmount and CurrentBalance start at 0."
        )
    elif isinstance(req, str) and "min 5 chars)." in req and "max 2000" not in req:
        ws["B2"].value = req.replace("min 5 chars).", "min 5 chars, max 2000).")

    for row in ws.iter_rows(min_row=10, max_row=1000):
        tc_id = row[0].value
        if not tc_id or not str(tc_id).startswith("CF_"):
            continue
        key = str(tc_id).strip()
        if key not in UPDATES:
            continue
        u = UPDATES[key]
        for col_letter, idx in (("B", 1), ("C", 2), ("D", 3), ("E", 4)):
            if col_letter in u:
                row[idx].value = u[col_letter]

    for row in ws.iter_rows(min_row=10, max_row=500):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and "signature sai" in v:
                cell.value = v.replace("signature sai", "invalid signature")

    _apply_reject_rows_style(ws)

    wb.save(XLSX)
    print(f"Updated: {XLSX}")


if __name__ == "__main__":
    main()
