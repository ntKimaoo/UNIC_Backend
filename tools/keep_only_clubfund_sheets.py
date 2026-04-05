from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def is_clubfund_sheet(name: str) -> bool:
    return ("ClubFund" in name) or ("Club.Fund" in name) or ("Club Fund" in name)


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    in_path = repo_root / "Unit Test.xlsx"
    out_path = repo_root / "Unit Test.clubfund.only.xlsx"

    import sys

    if len(sys.argv) >= 2:
        in_path = Path(sys.argv[1])
    if len(sys.argv) >= 3:
        out_path = Path(sys.argv[2])

    wb = load_workbook(in_path, data_only=False)
    keep = [s for s in wb.sheetnames if is_clubfund_sheet(s)]
    drop = [s for s in wb.sheetnames if s not in keep]

    if not keep:
        raise SystemExit("No ClubFund sheets found; nothing to keep.")

    # Delete non-ClubFund sheets
    for s in drop:
        wb.remove(wb[s])

    # Ensure an active sheet exists and is visible
    wb.active = 0
    for ws in wb.worksheets:
        ws.sheet_state = "visible"

    wb.save(out_path)
    print(f"kept={len(keep)} dropped={len(drop)} out={out_path}")


if __name__ == "__main__":
    main()

