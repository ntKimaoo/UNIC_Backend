from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def is_long_test_name(v: object) -> bool:
    if not isinstance(v, str):
        return False
    if len(v) <= 20:
        return False
    return ("_Should" in v) or v.endswith("Async") or ("Payment" in v) or ("TryComplete" in v)


def main() -> None:
    path = Path(__file__).resolve().parents[1] / "Unit Test.xlsx"
    wb = load_workbook(path, data_only=False)

    targets = [s for s in wb.sheetnames if s.startswith("S.ClubFund.")]
    changed = []

    for name in targets:
        ws = wb[name]

        # Most sheets follow a template where row 12 is a condition row containing markers (O),
        # but some ClubFund sheets accidentally have method/test names pasted into row 12 columns E+.
        # Fix rule (aligned with other sheets like S.Event.*):
        # - If row 13 already has 'O' in that column (meaning the next condition row uses that case),
        #   keep row 12 blank.
        # - Otherwise, row 12 should be 'O'.
        row = 12
        col_start = 5  # E
        col_end = ws.max_column

        local_changes = 0
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row, c)
            if not is_long_test_name(cell.value):
                continue

            below = ws.cell(row + 1, c).value
            new_val = "" if (below == "O") else "O"

            if cell.value != new_val:
                cell.value = new_val
                local_changes += 1

        if local_changes:
            changed.append((name, local_changes))

    wb.save(path)

    print(f"updated_sheets={len(changed)}/{len(targets)}")
    for name, count in changed:
        print(f"- {name}: {count} cells fixed")


if __name__ == "__main__":
    main()

