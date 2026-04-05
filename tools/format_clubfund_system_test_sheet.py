from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


def copy_cell_style(src: Cell, dst: Cell) -> None:
    # Copy full style object parts
    dst._style = src._style
    dst.number_format = src.number_format
    # Avoid assigning StyleProxy objects directly (can be unhashable in openpyxl internals)
    dst.font = src.font.copy()
    dst.fill = src.fill.copy()
    dst.border = src.border.copy()
    dst.alignment = src.alignment.copy()
    dst.protection = src.protection.copy()
    # Merged cells can be read-only for some attributes
    if getattr(dst, "comment", None) is not None:
        try:
            dst.comment = None
        except Exception:
            pass


def main() -> None:
    repo_root = Path(__file__).resolve().parents[1]
    path = repo_root / "System Test.xlsx"

    import sys

    if len(sys.argv) >= 2:
        path = Path(sys.argv[1])

    wb = load_workbook(path, data_only=False)
    if "Club Fund" not in wb.sheetnames:
        raise SystemExit("Sheet 'Club Fund' not found.")

    tmpl = wb["Template"]
    ws = wb["Club Fund"]

    # Use Template as authoritative for widths/heights & styles.
    # Column widths A..R
    for col_letter, dim in tmpl.column_dimensions.items():
        if col_letter in ws.column_dimensions:
            ws.column_dimensions[col_letter].width = dim.width
        else:
            ws.column_dimensions[col_letter].width = dim.width

    # Row heights 1..40 (enough to cover headers and early rows)
    for r in range(1, 41):
        if r in tmpl.row_dimensions and tmpl.row_dimensions[r].height is not None:
            ws.row_dimensions[r].height = tmpl.row_dimensions[r].height

    # Identify style rows in template
    # Row 10 is "Scenario 1- Mẫu"; row 11 is a sample TC.
    scenario_style_row = 10
    tc_style_row = 11

    max_col = tmpl.max_column

    # Apply scenario/tc styles to Club Fund rows from row 10 down to last populated row.
    last_row = ws.max_row
    for r in range(10, last_row + 1):
        a_val = ws.cell(r, 1).value
        if isinstance(a_val, str) and a_val.startswith("Scenario"):
            src_row = scenario_style_row
            ws.row_dimensions[r].height = tmpl.row_dimensions[scenario_style_row].height
        else:
            src_row = tc_style_row
            # Make testcase rows taller to avoid text overlap (match template if set, else 60)
            h = tmpl.row_dimensions[tc_style_row].height or 60
            ws.row_dimensions[r].height = h

        for c in range(1, max_col + 1):
            src = tmpl.cell(src_row, c)
            dst = ws.cell(r, c)
            copy_cell_style(src, dst)

    # Ensure wrap text for long text columns (B..F) like template.
    for r in range(10, last_row + 1):
        for c in range(2, 6 + 1):
            cell = ws.cell(r, c)
            if cell.alignment:
                cell.alignment = cell.alignment.copy(wrap_text=True)

    # Fix header workflow name / requirement row styles from template row 1..9.
    for r in range(1, 10):
        for c in range(1, max_col + 1):
            copy_cell_style(tmpl.cell(r, c), ws.cell(r, c))

    wb.save(path)
    print(f"formatted {path} sheet 'Club Fund'")


if __name__ == "__main__":
    main()

