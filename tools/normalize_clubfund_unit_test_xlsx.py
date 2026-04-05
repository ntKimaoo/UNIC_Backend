from __future__ import annotations

import datetime as dt
import re
from pathlib import Path

from openpyxl import load_workbook


UTCID_RE = re.compile(r"^UTCID\d+$")


def find_cell_with_value(ws, value: str, max_rows: int = 60, max_cols: int = 10):
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, min(ws.max_column, max_cols) + 1):
            if ws.cell(r, c).value == value:
                return r, c
    return None


def find_utcid_header_row(ws, max_rows: int = 40):
    for r in range(1, min(ws.max_row, max_rows) + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if isinstance(v, str) and UTCID_RE.match(v):
                return r
    return None


def get_total_test_cases(ws) -> int | None:
    pos = find_cell_with_value(ws, "Total Test Cases", max_rows=12, max_cols=40)
    if not pos:
        return None
    r, c = pos
    v = ws.cell(r + 1, c).value
    if isinstance(v, (int, float)) and v > 0:
        return int(v)
    return None


def normalize_pass_fail(v: object) -> object:
    if v == "Passed":
        return "P"
    if v == "Failed":
        return "F"
    return v


def column_has_any_data(ws, col: int, start_row: int, end_row: int) -> bool:
    for r in range(start_row, end_row + 1):
        if ws.cell(r, col).value not in (None, ""):
            return True
    return False


def guess_testcase_name_row(ws, utcid_cols: list[int], header_row: int) -> int | None:
    """
    Try to find the row that contains testcase names (e.g. Foo_Bar_WhenBaz).
    Many templates put these names around row 12.
    """
    best_row = None
    best_score = 0
    for r in range(1, min(ws.max_row, 40) + 1):
        if r == header_row:
            continue
        score = 0
        for c in utcid_cols:
            v = ws.cell(r, c).value
            if isinstance(v, str) and ("_" in v) and len(v) >= 8:
                score += 1
        if score > best_score:
            best_score = score
            best_row = r
    if best_score >= 1:
        return best_row
    return None


def is_testcase_like(v: object) -> bool:
    return isinstance(v, str) and ("_" in v) and len(v) >= 12


def fix_condition_row12_markers(ws, utcid_cols: list[int], row12: int = 12) -> bool:
    """
    Some sheets have method/test names pasted into the condition row.
    Rule: for each UTCID column at row12, ensure it's a marker:
    - If row13 has 'O' then row12 should be blank.
    - Otherwise row12 should be 'O'.
    """
    changed = False
    for c in utcid_cols:
        cell = ws.cell(row12, c)
        if is_testcase_like(cell.value):
            cell.value = None
            changed = True

        below = ws.cell(row12 + 1, c).value
        new_val = None if (below == "O") else "O"
        if cell.value in (None, "") and new_val == "O":
            cell.value = "O"
            changed = True
        elif cell.value == "O" and new_val is None:
            cell.value = None
            changed = True
        elif cell.value not in (None, "", "O") and is_testcase_like(cell.value):
            cell.value = new_val
            changed = True
    return changed


def normalize_clubid_inputs(ws) -> bool:
    """
    Many ClubFund sheets use ClubId=7 as a placeholder. In our actual tests, ClubId is typically 1.
    Keep 999 as the 'missing' case if present.
    """
    changed = False
    for r in range(1, min(ws.max_row, 30) + 1):
        if ws.cell(r, 2).value == "ClubId":
            # value cells usually in column D on next rows
            v1 = ws.cell(r + 1, 4).value
            if v1 == 7:
                ws.cell(r + 1, 4).value = 1
                changed = True
    return changed


def remove_999_clubid_case_if_single(ws) -> bool:
    """
    If Total Test Cases == 1, remove the 'missing club' (999) case from the input grid to avoid
    confusion. We keep the row structure but clear the 999 value and any markers on that row.
    """
    total = get_total_test_cases(ws)
    if total != 1:
        return False

    changed = False
    # Find the ClubId block.
    for r in range(1, min(ws.max_row, 40) + 1):
        if ws.cell(r, 2).value == "ClubId":
            # Common layout:
            # r+1 col D: valid clubId (e.g. 1)
            # r+2 col D: 999
            cand_row = r + 2
            if cand_row <= ws.max_row and ws.cell(cand_row, 4).value == 999:
                # Clear the 999 value and any markers across the row.
                ws.cell(cand_row, 4).value = None
                for c in range(5, ws.max_column + 1):
                    if ws.cell(cand_row, c).value == "O":
                        ws.cell(cand_row, c).value = None
                changed = True
            break
    return changed


def delete_trailing_blank_columns(ws) -> bool:
    """
    Physically delete columns at the far right that are completely blank across the visible template
    (top 1..25 rows). This makes the sheet less wide without touching meaningful cells.
    """
    max_check_row = min(ws.max_row, 25)
    last_nonempty = 1
    for c in range(1, ws.max_column + 1):
        for r in range(1, max_check_row + 1):
            if ws.cell(r, c).value not in (None, ""):
                last_nonempty = c
                break

    if last_nonempty >= ws.max_column:
        return False

    # Delete from the end down to last_nonempty+1
    delete_count = ws.max_column - last_nonempty
    ws.delete_cols(last_nonempty + 1, delete_count)
    return True


def main() -> None:
    today = dt.date(2026, 3, 30)  # per user's "ngày hôm nay"
    repo_root = Path(__file__).resolve().parents[1]
    in_path = repo_root / "Unit Test.xlsx"
    out_path = repo_root / "Unit Test.xlsx"

    import sys

    if len(sys.argv) >= 2:
        in_path = Path(sys.argv[1])
    if len(sys.argv) >= 3:
        out_path = Path(sys.argv[2])

    wb = load_workbook(in_path, data_only=False)

    targets = [s for s in wb.sheetnames if "ClubFund" in s]
    changed_sheets: list[str] = []

    for name in targets:
        ws = wb[name]
        header_row = find_utcid_header_row(ws)
        if header_row is None:
            continue

        # Identify UTCID columns.
        utcid_cols: list[int] = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if isinstance(v, str) and UTCID_RE.match(v):
                utcid_cols.append(c)
        if not utcid_cols:
            continue

        changed = False

        # Make conditions more consistent with actual inputs.
        if normalize_clubid_inputs(ws):
            changed = True

        if remove_999_clubid_case_if_single(ws):
            changed = True

        # Locate "Passed/Failed" and "Executed Date" rows (labels usually in col B).
        pf_pos = find_cell_with_value(ws, "Passed/Failed", max_rows=60, max_cols=12)
        ed_pos = find_cell_with_value(ws, "Executed Date", max_rows=60, max_cols=12)

        if pf_pos:
            pf_row, pf_col = pf_pos
            for c in utcid_cols:
                old = ws.cell(pf_row, c).value
                new = normalize_pass_fail(old)
                if new != old:
                    ws.cell(pf_row, c).value = new
                    changed = True

        # Fill Executed Date for any test case that has P/F.
        if pf_pos and ed_pos:
            pf_row, _ = pf_pos
            ed_row, _ = ed_pos
            for c in utcid_cols:
                v = ws.cell(pf_row, c).value
                if v in ("P", "F"):
                    cell = ws.cell(ed_row, c)
                    cell.value = today
                    cell.number_format = "dd/mm"
                    changed = True

        # Delete trailing unused UTCID columns.
        # Define data rows as from Condition block start to Effect ID row (if present), else up to max_row.
        cond_pos = find_cell_with_value(ws, "Condition", max_rows=80, max_cols=6)
        eff_pos = find_cell_with_value(ws, "Effect ID", max_rows=120, max_cols=12)
        start_row = cond_pos[0] if cond_pos else header_row
        end_row = eff_pos[0] if eff_pos else ws.max_row

        # Fix row12 condition markers (and remove pasted testcase names if any).
        if fix_condition_row12_markers(ws, utcid_cols, row12=12):
            changed = True

        # Prefer trimming based on "Total Test Cases" if present.
        total_cases = get_total_test_cases(ws)
        last_used_col = None
        if total_cases is not None and total_cases >= 1:
            keep = utcid_cols[:total_cases]
            if keep:
                last_used_col = keep[-1]
        else:
            # Otherwise, trim based on the last non-empty testcase-name cell, if we can detect it.
            tc_row = guess_testcase_name_row(ws, utcid_cols, header_row)
            if tc_row is not None:
                for c in utcid_cols:
                    v = ws.cell(tc_row, c).value
                    if v not in (None, ""):
                        last_used_col = c
            else:
                for c in utcid_cols:
                    if column_has_any_data(ws, c, start_row, end_row):
                        last_used_col = c

        if last_used_col is not None:
            # Instead of deleting whole worksheet columns (which would also remove summary cells
            # like "Total Test Cases"), clear the unused UTCID columns within the test matrix.
            cols_to_clear = [c for c in utcid_cols if c > last_used_col]
            if cols_to_clear:
                clear_start = header_row
                clear_end = end_row
                for c in cols_to_clear:
                    for r in range(clear_start, clear_end + 1):
                        ws.cell(r, c).value = None
                changed = True

        if delete_trailing_blank_columns(ws):
            changed = True

        if changed:
            changed_sheets.append(name)

    wb.save(out_path)
    print(f"normalized_sheets={len(changed_sheets)}/{len(targets)}")
    for s in changed_sheets[:50]:
        print(f"- {s}")


if __name__ == "__main__":
    main()

