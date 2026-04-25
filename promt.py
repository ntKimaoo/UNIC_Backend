import openpyxl
from copy import copy
from openpyxl.utils import get_column_letter

SRC = r'c:\DATN\UniClub_Backend\unittest.xlsx'
DST = r'c:\DATN\UniClub_Backend\unittest_report.xlsx'

wb = openpyxl.load_workbook(SRC)
template_ws = wb[wb.sheetnames[0]]

def copy_sheet(wb, template_ws, new_name):
    """Copy template sheet with all formatting, return new sheet."""
    new_ws = wb.copy_worksheet(template_ws)
    new_ws.title = new_name[:31]
    return new_ws

def clear_data(ws):
    """Clear only the data cells, keep structure/labels."""
    # Clear header data (C1=ModuleName, K1=Method, C2=DevName, C3=Description)
    ws['C1'] = None; ws['K1'] = None; ws['C2'] = None; ws['C3'] = None
    # Clear summary counts (A5, C5, E5, K5, L5, M5, N5)
    ws['A5'] = None; ws['C5'] = None; ws['E5'] = None
    ws['K5'] = None; ws['L5'] = None; ws['M5'] = None; ws['N5'] = None
    # Clear UTCID row 7 (E7:S7)
    for c in range(5, 20):
        ws.cell(row=7, column=c).value = None
    # Clear input data rows 9-13
    ws['D9'] = None
    ws['D11'] = None; ws['E11'] = None
    ws['D13'] = None; ws['F13'] = None
    # Clear B10 input label
    ws['B10'] = None
    # Clear return data rows 15-16
    ws['D15'] = None; ws['E15'] = None
    ws['D16'] = None; ws['F16'] = None
    # Clear exception rows 18-19
    ws['D18'] = None; ws['D19'] = None; ws['F19'] = None
    # Clear log rows 22-23
    ws['D22'] = None; ws['E22'] = None
    ws['D23'] = None; ws['F23'] = None
    # Clear Result row 24-27 (E-S columns)
    for r in range(24, 28):
        for c in range(5, 20):
            ws.cell(row=r, column=c).value = None

def fill_sheet(ws, module, method, layer, desc, test_cases):
    """
    test_cases: list of dicts:
      preconditions: list of str
      input_label: str (e.g. 'LoginRequestDto')
      inputs: list of (value_str, col_marks) where col_marks = list of utcid_index (0-based)
      returns: list of (value_str, col_marks)
      exceptions: list of (value_str, col_marks)
      logs: list of (value_str, col_marks)
      types: list of str (N/A/B per test case)
      results: list of str (P/F per test case)
    """
    num_tc = len(test_cases['types'])
    
    # Header
    ws['C1'] = f'{layer}.{module}'
    ws['K1'] = method
    ws['C2'] = 'Developer'
    ws['C3'] = desc
    
    # Summary
    p = test_cases['results'].count('P')
    f = test_cases['results'].count('F')
    n = test_cases['types'].count('N')
    a = test_cases['types'].count('A')
    b = test_cases['types'].count('B')
    ws['A5'] = p; ws['C5'] = f; ws['E5'] = 0
    ws['K5'] = n; ws['L5'] = a; ws['M5'] = b; ws['N5'] = num_tc
    
    # UTCID headers (row 7)
    for i in range(num_tc):
        cell = ws.cell(row=7, column=5+i)
        cell.value = f'UTCID{i+1:02d}'
        # Copy style from template E7
        tc = template_ws.cell(row=7, column=5)
        cell.font = copy(tc.font)
        cell.fill = copy(tc.fill)
        cell.alignment = copy(tc.alignment)
        cell.border = copy(tc.border)
    
    # Preconditions (row 9+)
    row = 9
    for pc in test_cases.get('preconditions', []):
        ws.cell(row=row, column=4).value = pc[0]
        for ci in pc[1]:
            ws.cell(row=row, column=5+ci).value = 'O'
            ws.cell(row=row, column=5+ci).font = copy(template_ws['E11'].font)
            ws.cell(row=row, column=5+ci).alignment = copy(template_ws['E11'].alignment)
        row += 1
    
    # Input label
    if test_cases.get('input_label'):
        ws.cell(row=row, column=2).value = test_cases['input_label']
        ws.cell(row=row, column=2).font = copy(template_ws['B10'].font)
        row += 1
    
    # Input values
    for inp in test_cases.get('inputs', []):
        ws.cell(row=row, column=4).value = inp[0]
        for ci in inp[1]:
            ws.cell(row=row, column=5+ci).value = 'O'
            ws.cell(row=row, column=5+ci).font = copy(template_ws['E11'].font)
            ws.cell(row=row, column=5+ci).alignment = copy(template_ws['E11'].alignment)
        row += 1
    
    # Confirm - Return (reuse rows 14+)
    # Row 14 is "Confirm" + "Return" header - already in template
    # We need to place returns starting at row 15
    row = 15
    for ret in test_cases.get('returns', []):
        ws.cell(row=row, column=4).value = ret[0]
        for ci in ret[1]:
            ws.cell(row=row, column=5+ci).value = 'O'
            ws.cell(row=row, column=5+ci).font = copy(template_ws['E11'].font)
            ws.cell(row=row, column=5+ci).alignment = copy(template_ws['E11'].alignment)
        row += 1
    
    # Exception (row 17 is header)
    row = 18
    for exc in test_cases.get('exceptions', []):
        ws.cell(row=row, column=4).value = exc[0]
        for ci in exc[1]:
            ws.cell(row=row, column=5+ci).value = 'O'
            ws.cell(row=row, column=5+ci).font = copy(template_ws['E11'].font)
            ws.cell(row=row, column=5+ci).alignment = copy(template_ws['E11'].alignment)
        row += 1
    
    # Log message (row 20 is header)
    row = 22
    for log in test_cases.get('logs', []):
        ws.cell(row=row, column=4).value = log[0]
        for ci in log[1]:
            ws.cell(row=row, column=5+ci).value = 'O'
            ws.cell(row=row, column=5+ci).font = copy(template_ws['E11'].font)
            ws.cell(row=row, column=5+ci).alignment = copy(template_ws['E11'].alignment)
        row += 1
    
    # Result rows 24-27
    for i, t in enumerate(test_cases['types']):
        ws.cell(row=24, column=5+i).value = t
        ws.cell(row=24, column=5+i).alignment = copy(template_ws['E24'].alignment)
        ws.cell(row=24, column=5+i).font = copy(template_ws['E24'].font)
    for i, r in enumerate(test_cases['results']):
        ws.cell(row=25, column=5+i).value = r
        ws.cell(row=25, column=5+i).alignment = copy(template_ws['E25'].alignment)
        ws.cell(row=25, column=5+i).font = copy(template_ws['E25'].font)
    for i in range(num_tc):
        ws.cell(row=26, column=5+i).value = '03/23'
        ws.cell(row=26, column=5+i).alignment = copy(template_ws['E26'].alignment)
        ws.cell(row=26, column=5+i).font = copy(template_ws['E26'].font)

def make(name, module, method, layer, desc, data):
    ws = copy_sheet(wb, template_ws, name)
    clear_data(ws)
    fill_sheet(ws, module, method, layer, desc, data)

# Helper
def d(preconds, input_label, inputs, returns, exceptions, logs, types, results):
    return {
        'preconditions': preconds, 'input_label': input_label,
        'inputs': inputs, 'returns': returns, 'exceptions': exceptions,
        'logs': logs, 'types': types, 'results': results
    }

ALL = list(range(20))  # shortcut: all test cases

# ===== CONTROLLER TESTS =====

make('C.Auth.Login', 'AuthController', 'Login', 'Controller',
    'Verify login returns correct HTTP status codes',
    d([('Can connect with server', [0,1])], 'LoginRequestDto',
      [('{test@test.com, password}', [0]), ('{test@test.com, wrong}', [1])],
      [('{token, refresh}', [0]), ('null', [1])],
      [('UnauthorizedException', [1])],
      [('"Login successful"', [0]), ('"Invalid credentials"', [1])],
      ['N','A'], ['P','P']))

make('C.Auth.Register', 'AuthController', 'Register', 'Controller',
    'Verify register returns correct HTTP status codes',
    d([('Can connect with server', [0,1,2])], 'RegisterRequestDto',
      [('{User, new@test.com, pwd}', [0]), ('{User, exists@test.com, pwd}', [1]), ('{User, test@test.com, pwd}', [2])],
      [('201 CreatedAtActionResult', [0]), ('400 BadRequestObjectResult', [1,2])],
      [('InvalidOperationException', [1])],
      [],
      ['N','A','A'], ['P','P','P']))

make('C.Auth.VerifyEmail', 'AuthController', 'VerifyEmail', 'Controller',
    'Verify email verification endpoint',
    d([('Can connect with server', [0,1])], 'VerifyEmailRequestDto',
      [('{test@test.com, valid-token}', [0]), ('{test@test.com, bad-token}', [1])],
      [('200 OkObjectResult', [0]), ('400 BadRequestObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Auth.RefreshToken', 'AuthController', 'RefreshToken', 'Controller',
    'Verify refresh token endpoint',
    d([('Can connect with server', [0,1])], 'RefreshTokenRequestDto',
      [('RefreshToken=valid-token', [0]), ('RefreshToken=invalid-token', [1])],
      [('200 OkObjectResult', [0]), ('401 UnauthorizedObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Auth.ForgotPwd', 'AuthController', 'ForgotPassword', 'Controller',
    'Verify forgot password endpoint (anti-enumeration)',
    d([('Can connect with server', [0])], 'ForgotPasswordRequestDto',
      [('Email=test@test.com', [0])],
      [('200 OkObjectResult', [0])],
      [], [],
      ['N'], ['P']))

make('C.Auth.ResetPwd', 'AuthController', 'ResetPassword', 'Controller',
    'Verify reset password endpoint',
    d([('Can connect with server', [0,1])], 'ResetPasswordRequestDto',
      [('{test@test.com, valid-token, new}', [0]), ('{test@test.com, expired, new}', [1])],
      [('200 OkObjectResult', [0]), ('400 BadRequestObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Auth.RevokeToken', 'AuthController', 'RevokeToken', 'Controller',
    'Verify logout (revoke token) endpoint',
    d([('Can connect with server', [0,1])], 'RefreshTokenRequestDto',
      [('RefreshToken=valid-token', [0]), ('RefreshToken=invalid', [1])],
      [('200 OkObjectResult', [0]), ('400 BadRequestObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Auth.LogoutAll', 'AuthController', 'LogoutAllDevices', 'Controller',
    'Verify logout all devices endpoint',
    d([('Can connect with server', [0,1]), ('User authenticated', [0])], 'Claims',
      [('Valid userId+email', [0]), ('Empty ClaimsIdentity', [1])],
      [('200 OkObjectResult', [0]), ('401 UnauthorizedObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Auth.ChangePwd', 'AuthController', 'ChangePassword', 'Controller',
    'Verify change password endpoint',
    d([('Can connect with server', [0,1]), ('User authenticated', [0,1])], 'ChangePasswordRequestDto',
      [('{old, new}', [0]), ('{wrong, new}', [1])],
      [('200 OkObjectResult', [0]), ('400 BadRequestObjectResult', [1])],
      [('InvalidOperationException', [1])], [],
      ['N','A'], ['P','P']))

make('C.Auth.ResendVerify', 'AuthController', 'ResendVerification', 'Controller',
    'Verify resend verification email endpoint',
    d([('Can connect with server', [0,1])], 'ForgotPasswordRequestDto',
      [('Email=test@test.com', [0]), ('Email=missing@test.com', [1])],
      [('200 OkObjectResult', [0]), ('400 BadRequestObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Auth.GetProfile', 'AuthController', 'GetCurrentUser', 'Controller',
    'Verify get current user profile endpoint',
    d([('Can connect with server', [0,1]), ('User authenticated', [0])], 'Claims',
      [('Valid userId+email+name', [0]), ('Empty ClaimsIdentity', [1])],
      [('200 OkObjectResult', [0]), ('401 UnauthorizedObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

# --- UsersController ---
make('C.Users.GetAll', 'UsersController', 'GetAll', 'Controller',
    'Verify paged user list endpoint',
    d([('Can connect with server', [0,1,2])], 'Query params',
      [('page=1, pageSize=10', [0]), ('page=0, pageSize=10', [1]), ('page=1, pageSize=200', [2])],
      [('200 OkObjectResult', [0]), ('400 BadRequestObjectResult', [1,2])],
      [], [],
      ['N','A','B'], ['P','P','P']))

make('C.Users.GetById', 'UsersController', 'GetById', 'Controller',
    'Verify get user by ID endpoint',
    d([('Can connect with server', [0,1])], 'Path param',
      [('userId=Valid GUID', [0]), ('userId=Random GUID', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Users.Create', 'UsersController', 'Create', 'Controller',
    'Verify create user endpoint',
    d([('Can connect with server', [0,1])], 'CreateUserDto',
      [('{Test, new@test.com, pwd}', [0]), ('{Test, exists@test.com, pwd}', [1])],
      [('201 CreatedAtActionResult', [0]), ('400 BadRequestObjectResult', [1])],
      [('Exception("Email already exists.")', [1])], [],
      ['N','A'], ['P','P']))

make('C.Users.Update', 'UsersController', 'Update', 'Controller',
    'Verify update user endpoint',
    d([('Can connect with server', [0,1,2])], 'UpdateUserDto',
      [('FullName=New Name', [0,1]), ('StudentId=exists', [2])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1]), ('400 BadRequestObjectResult', [2])],
      [('Exception("Student ID already exists.")', [2])], [],
      ['N','A','A'], ['P','P','P']))

make('C.Users.Delete', 'UsersController', 'Delete', 'Controller',
    'Verify delete user endpoint',
    d([('Can connect with server', [0,1])], 'Path param',
      [('userId=Valid GUID', [0]), ('userId=Random GUID', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Users.GetAllClub', 'UsersController', 'GetAllClub', 'Controller',
    'Verify get all clubs by user endpoint',
    d([('Can connect with server', [0,1,2]), ('User exists', [0,2]), ('User has clubs', [0])], 'Path param',
      [('userId=Valid GUID', [0,2]), ('userId=Random GUID', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1,2])],
      [], [],
      ['N','A','A'], ['P','P','P']))

make('C.Users.ManagedClubs', 'UsersController', 'GetManagedClubs', 'Controller',
    'Verify get managed clubs endpoint',
    d([('Can connect with server', [0])], 'Path param',
      [('userId=Valid GUID', [0])],
      [('200 OkObjectResult', [0])],
      [], [],
      ['N'], ['P']))

# --- ClubPostController ---
make('C.Post.GetAll', 'ClubPostController', 'GetAll', 'Controller',
    'Verify get all posts by club endpoint',
    d([('Can connect with server', [0])], 'Path param',
      [('clubId=1', [0])],
      [('200 OkObjectResult', [0])],
      [], [],
      ['N'], ['P']))

make('C.Post.GetById', 'ClubPostController', 'GetById', 'Controller',
    'Verify get post by ID endpoint',
    d([('Can connect with server', [0,1])], 'Path params',
      [('clubId=1, postId=1', [0]), ('clubId=1, postId=99', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Post.GetByUser', 'ClubPostController', 'GetByUserId', 'Controller',
    'Verify get posts by user ID endpoint',
    d([('Can connect with server', [0])], 'Path params',
      [('clubId=1, userId=GUID', [0])],
      [('200 OkObjectResult', [0])],
      [], [],
      ['N'], ['P']))

make('C.Post.Create', 'ClubPostController', 'Create', 'Controller',
    'Verify create post endpoint',
    d([('Can connect with server', [0,1])], 'CreateClubPostDto',
      [('{ClubId=1, Title=New Post}', [0]), ('{ClubId=1, Title=Bad}', [1])],
      [('201 CreatedAtActionResult', [0]), ('400 BadRequestObjectResult', [1])],
      [('Exception("Creation failed")', [1])], [],
      ['N','A'], ['P','P']))

make('C.Post.Update', 'ClubPostController', 'Update', 'Controller',
    'Verify update post endpoint',
    d([('Can connect with server', [0,1])], 'UpdateClubPostDto',
      [('postId=1, Title=Updated', [0]), ('postId=99, Title=Missing', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Post.Delete', 'ClubPostController', 'Delete', 'Controller',
    'Verify delete post endpoint',
    d([('Can connect with server', [0,1])], 'Path params',
      [('clubId=1, postId=1', [0]), ('clubId=1, postId=99', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Post.UploadImg', 'ClubPostController', 'UploadEditorImage', 'Controller',
    'Verify upload editor image endpoint',
    d([('Can connect with server', [0,1])], 'IFormFile',
      [('file=null', [0]), ('file=image.png (valid)', [1])],
      [('400 BadRequestObjectResult', [0]), ('200 OkObjectResult', [1])],
      [], [],
      ['A','N'], ['P','P']))

# --- InterviewsController ---
make('C.Intv.Create', 'InterviewsController', 'Create', 'Controller',
    'Verify create interview schedule endpoint',
    d([('Can connect with server', [0,1])], 'CreateInterviewScheduleDto',
      [('Title=Test Interview', [0]), ('Title=Bad', [1])],
      [('201 CreatedAtActionResult', [0]), ('400 BadRequestObjectResult', [1])],
      [('Exception("Invalid data")', [1])], [],
      ['N','A'], ['P','P']))

make('C.Intv.GetAll', 'InterviewsController', 'GetAll', 'Controller',
    'Verify get all schedules endpoint',
    d([('Can connect with server', [0])], 'Query params',
      [('filters=null', [0])],
      [('200 OkObjectResult', [0])],
      [], [],
      ['N'], ['P']))

make('C.Intv.GetById', 'InterviewsController', 'GetById', 'Controller',
    'Verify get schedule by ID endpoint',
    d([('Can connect with server', [0,1])], 'Path param',
      [('scheduleId=1', [0]), ('scheduleId=99', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Intv.Update', 'InterviewsController', 'Update', 'Controller',
    'Verify update schedule endpoint',
    d([('Can connect with server', [0,1,2])], 'UpdateInterviewScheduleDto',
      [('id=1, Title=Updated', [0]), ('id=99, Title=Missing', [1]), ('id=1, Title=Bad', [2])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1]), ('400 BadRequestObjectResult', [2])],
      [('Exception("Error")', [2])], [],
      ['N','A','A'], ['P','P','P']))

make('C.Intv.UpdStatus', 'InterviewsController', 'UpdateStatus', 'Controller',
    'Verify update interview status endpoint',
    d([('Can connect with server', [0,1,2])], 'UpdateInterviewStatusDto',
      [('id=1, Status=Confirmed', [0]), ('id=99, Status=Confirmed', [1]), ('id=1, Status=Invalid', [2])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1]), ('400 BadRequestObjectResult', [2])],
      [('ArgumentException("Invalid status")', [2])], [],
      ['N','A','A'], ['P','P','P']))

make('C.Intv.Delete', 'InterviewsController', 'Delete', 'Controller',
    'Verify delete schedule endpoint',
    d([('Can connect with server', [0,1,2])], 'Path param',
      [('scheduleId=1', [0]), ('scheduleId=99', [1]), ('scheduleId=1 (cannot delete)', [2])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1]), ('400 BadRequestObjectResult', [2])],
      [('InvalidOperationException', [2])], [],
      ['N','A','A'], ['P','P','P']))

make('C.Intv.AssignIntv', 'InterviewsController', 'AssignInterviewers', 'Controller',
    'Verify assign interviewers endpoint',
    d([('Can connect with server', [0,1])], 'AssignInterviewersDto',
      [('id=1, 1 valid interviewer', [0]), ('id=99, empty list', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [('KeyNotFoundException', [1])], [],
      ['N','A'], ['P','P']))

make('C.Intv.GetAssign', 'InterviewsController', 'GetAssignments', 'Controller',
    'Verify get assignments endpoint',
    d([('Can connect with server', [0])], 'Path param',
      [('scheduleId=1', [0])],
      [('200 OkObjectResult', [0])],
      [], [],
      ['N'], ['P']))

make('C.Intv.RmAssign', 'InterviewsController', 'RemoveAssignment', 'Controller',
    'Verify remove assignment endpoint',
    d([('Can connect with server', [0,1])], 'Path params',
      [('scheduleId=1, assignmentId=2', [0]), ('scheduleId=1, assignmentId=99', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Intv.ConfAssign', 'InterviewsController', 'ConfirmAssignment', 'Controller',
    'Verify confirm assignment endpoint',
    d([('Can connect with server', [0,1])], 'Path params',
      [('scheduleId=1, assignmentId=2', [0]), ('scheduleId=1, assignmentId=99', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Intv.GetRoom', 'InterviewsController', 'GetRoom', 'Controller',
    'Verify get room by schedule endpoint',
    d([('Can connect with server', [0,1])], 'Path param',
      [('scheduleId=1', [0]), ('scheduleId=99', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Intv.Feedback', 'InterviewsController', 'SubmitFeedback', 'Controller',
    'Verify submit feedback endpoint',
    d([('Can connect with server', [0,1])], 'SubmitFeedbackDto',
      [('id=1, asgId=2, Pass, Score=8', [0]), ('id=1, asgId=99, Pass', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Intv.FbSummary', 'InterviewsController', 'GetFeedbackSummary', 'Controller',
    'Verify get feedback summary endpoint',
    d([('Can connect with server', [0,1])], 'Path param',
      [('scheduleId=1', [0]), ('scheduleId=99', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

# --- RoomsController ---
make('C.Room.Join', 'RoomsController', 'JoinRoom', 'Controller',
    'Verify join room endpoint',
    d([('Can connect with server', [0,1,2]), ('Room open', [0])], 'JoinRoomDto',
      [('roomCode=abc-1234, User', [0]), ('roomCode=missing', [1]), ('roomCode=full', [2])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1]), ('400 BadRequestObjectResult', [2])],
      [('KeyNotFoundException', [1]), ('InvalidOperationException("Room da day")', [2])], [],
      ['N','A','A'], ['P','P','P']))

make('C.Room.Leave', 'RoomsController', 'LeaveRoom', 'Controller',
    'Verify leave room endpoint',
    d([('Can connect with server', [0,1])], 'LeaveRoomDto',
      [('roomCode=abc-1234', [0]), ('roomCode=missing', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

make('C.Room.GetParts', 'RoomsController', 'GetParticipants', 'Controller',
    'Verify get participants endpoint',
    d([('Can connect with server', [0,1])], 'Path param',
      [('roomCode=abc-1234', [0]), ('roomCode=missing', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [('KeyNotFoundException', [1])], [],
      ['N','A'], ['P','P']))

make('C.Room.GetEvents', 'RoomsController', 'GetEvents', 'Controller',
    'Verify get events endpoint',
    d([('Can connect with server', [0,1])], 'Path param',
      [('roomCode=abc-1234', [0]), ('roomCode=missing', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [('KeyNotFoundException', [1])], [],
      ['N','A'], ['P','P']))

make('C.Room.Close', 'RoomsController', 'CloseRoom', 'Controller',
    'Verify close room endpoint',
    d([('Can connect with server', [0,1])], 'Path param',
      [('roomCode=abc-1234', [0]), ('roomCode=missing', [1])],
      [('200 OkObjectResult', [0]), ('404 NotFoundObjectResult', [1])],
      [], [],
      ['N','A'], ['P','P']))

# ===== SERVICE TESTS =====
make('S.Auth.Login', 'AuthService', 'LoginAsync', 'Service',
    'Verify login business logic',
    d([('Repository available', [0,1,2])], 'LoginRequestDto + User state',
      [('User not found', [0]), ('User Pending', [1]), ('User Active, valid pwd', [2])],
      [('null', [0,1]), ('LoginResponseDto (tokens+user)', [2])],
      [], [],
      ['A','A','N'], ['P','P','P']))

make('S.Auth.Refresh', 'AuthService', 'RefreshTokenAsync', 'Service',
    'Verify refresh token business logic',
    d([('Repository available', [0,1,2,3])], 'RefreshToken + Token/User state',
      [('Token not found', [0]), ('Token expired', [1]), ('User Pending', [2]), ('All valid', [3])],
      [('null', [0,1,2]), ('LoginResponseDto (new tokens)', [3])],
      [], [],
      ['A','A','A','N'], ['P','P','P','P']))

make('S.Auth.Revoke', 'AuthService', 'RevokeTokenAsync', 'Service',
    'Verify revoke token business logic',
    d([('Repository available', [0,1])], 'Token state',
      [('Token not found/revoked', [0]), ('Token valid, not revoked', [1])],
      [('false', [0]), ('true (IsRevoked=true)', [1])],
      [], [],
      ['A','N'], ['P','P']))

make('S.Auth.LogoutAll', 'AuthService', 'LogoutAllDevicesAsync', 'Service',
    'Verify logout all devices logic',
    d([('Repository available', [0])], 'userId',
      [('userId=Valid GUID', [0])],
      [('true (RevokeAll called)', [0])],
      [], [],
      ['N'], ['P']))

make('S.Auth.Register', 'AuthService', 'RegisterAsync', 'Service',
    'Verify register business logic',
    d([('Repository available', [0,1])], 'RegisterRequestDto',
      [('Email exists=Yes', [0]), ('Email exists=No', [1])],
      [('UserInfoDto', [1])],
      [('InvalidOperationException', [0])], [],
      ['A','N'], ['P','P']))

make('S.Auth.ChangePwd', 'AuthService', 'ChangePasswordAsync', 'Service',
    'Verify change password business logic',
    d([('Repository available', [0,1,2])], 'ChangePasswordRequestDto + User state',
      [('User not found', [0]), ('Wrong current pwd', [1]), ('Correct pwd', [2])],
      [('false', [0]), ('true (tokens revoked)', [2])],
      [('InvalidOperationException', [1])], [],
      ['A','A','N'], ['P','P','P']))

make('S.Auth.ForgotPwd', 'AuthService', 'ForgotPasswordAsync', 'Service',
    'Verify forgot password logic (anti-enumeration)',
    d([('Repository available', [0,1])], 'ForgotPasswordRequestDto',
      [('User not found', [0]), ('User found', [1])],
      [('true (anti-enumeration)', [0,1])],
      [], [],
      ['N','N'], ['P','P']))

make('S.Auth.ResetPwd', 'AuthService', 'ResetPasswordAsync', 'Service',
    'Verify reset password business logic',
    d([('Repository available', [0,1])], 'ResetPasswordRequestDto',
      [('Token not found', [0]), ('Token valid', [1])],
      [('false', [0]), ('true (pwd updated, tokens revoked)', [1])],
      [], [],
      ['A','N'], ['P','P']))

make('S.Auth.VerifyEmail', 'AuthService', 'VerifyEmailAsync', 'Service',
    'Verify email verification business logic',
    d([('Repository available', [0,1])], 'VerifyEmailRequestDto',
      [('Token not found', [0]), ('Token valid', [1])],
      [('false', [0]), ('true (Status->Active)', [1])],
      [], [],
      ['A','N'], ['P','P']))

make('S.Auth.ResendVerif', 'AuthService', 'ResendVerificationEmailAsync', 'Service',
    'Verify resend verification email logic',
    d([('Repository available', [0,1,2])], 'Email + User state',
      [('User not found', [0]), ('User Active', [1]), ('User Pending', [2])],
      [('false', [0,1]), ('true (new token created)', [2])],
      [], [],
      ['A','A','N'], ['P','P','P']))

# --- UserService ---
make('S.User.GetAll', 'UserService', 'GetAllUsersAsync', 'Service',
    'Verify get all users logic',
    d([('Repository available', [0])], '(no params)',
      [('2 users in DB', [0])],
      [('List of 2 UserResponseDto', [0])],
      [], [],
      ['N'], ['P']))

make('S.User.GetById', 'UserService', 'GetUserByIdAsync', 'Service',
    'Verify get user by ID logic',
    d([('Repository available', [0,1])], 'userId',
      [('User not found', [0]), ('User found', [1])],
      [('null', [0]), ('UserResponseDto', [1])],
      [], [],
      ['A','N'], ['P','P']))

make('S.User.Create', 'UserService', 'CreateUserAsync', 'Service',
    'Verify create user logic with duplicate checks',
    d([('Repository available', [0,1,2])], 'CreateUserDto',
      [('Email exists', [0]), ('StudentId exists', [1]), ('No conflicts', [2])],
      [('UserResponseDto', [2])],
      [('Exception("Email already exists.")', [0]), ('Exception("Student ID already exists.")', [1])], [],
      ['A','A','N'], ['P','P','P']))

make('S.User.Update', 'UserService', 'UpdateUserAsync', 'Service',
    'Verify update user logic with duplicate checks',
    d([('Repository available', [0,1,2])], 'UpdateUserDto',
      [('User not found', [0]), ('StudentId conflict', [1]), ('Valid update', [2])],
      [('false', [0]), ('true (fields updated)', [2])],
      [('Exception("Student ID already exists.")', [1])], [],
      ['A','A','N'], ['P','P','P']))

make('S.User.Delete', 'UserService', 'DeleteUserAsync', 'Service',
    'Verify delete user logic',
    d([('Repository available', [0])], 'userId',
      [('userId=Valid GUID', [0])],
      [('true', [0])],
      [], [],
      ['N'], ['P']))

# --- ClubPostService ---
make('S.Post.GetById', 'ClubPostService', 'GetByIdAsync', 'Service',
    'Verify get post by ID with DTO mapping',
    d([('Repository available', [0,1])], 'postId',
      [('Post not found', [0]), ('Post found', [1])],
      [('null', [0]), ('ClubPostResponseDto (mapped)', [1])],
      [], [],
      ['A','N'], ['P','P']))

make('S.Post.GetAll', 'ClubPostService', 'GetAllAsync', 'Service',
    'Verify get all posts logic',
    d([('Repository available', [0])], '(no params)',
      [('1 post in DB', [0])],
      [('List of DTOs', [0])],
      [], [],
      ['N'], ['P']))

make('S.Post.GetByClub', 'ClubPostService', 'GetByClubIdAsync', 'Service',
    'Verify get posts by club ID logic',
    d([('Repository available', [0])], 'clubId',
      [('clubId=2', [0])],
      [('List with ClubId=2', [0])],
      [], [],
      ['N'], ['P']))

make('S.Post.GetByUser', 'ClubPostService', 'GetByUserIdAsync', 'Service',
    'Verify get posts by user ID logic',
    d([('Repository available', [0])], 'userId',
      [('userId=Valid GUID', [0])],
      [('List with matching userId', [0])],
      [], [],
      ['N'], ['P']))

make('S.Post.Create', 'ClubPostService', 'CreateAsync', 'Service',
    'Verify create post logic (PUBLISHED vs PENDING)',
    d([('Repository available', [0,1])], 'CreateClubPostDto + IFormFile',
      [('image=null', [0]), ('image=test.png', [1])],
      [('Status=PUBLISHED', [0]), ('Status=PENDING', [1])],
      [], [],
      ['N','N'], ['P','P']))

make('S.Post.Update', 'ClubPostService', 'UpdateAsync', 'Service',
    'Verify update post logic with image handling',
    d([('Repository available', [0,1,2,3])], 'UpdateClubPostDto + IFormFile',
      [('Post not found', [0]), ('No image, fields updated', [1]), ('With image', [2]), ('Repo update fails', [3])],
      [('null', [0,3]), ('DTO (Status=DRAFT)', [1]), ('DTO (Status=PENDING)', [2])],
      [], [],
      ['A','N','N','A'], ['P','P','P','P']))

make('S.Post.Delete', 'ClubPostService', 'DeleteAsync', 'Service',
    'Verify delete post logic',
    d([('Repository available', [0])], 'postId',
      [('postId=1', [0])],
      [('true', [0])],
      [], [],
      ['N'], ['P']))

# --- InterviewService ---
make('S.Intv.Create', 'InterviewService', 'CreateScheduleAsync', 'Service',
    'Verify create schedule with room and assignments',
    d([('Repository available', [0])], 'CreateInterviewScheduleDto',
      [('Title=Test, 1 interviewer', [0])],
      [('InterviewScheduleResponseDto', [0])],
      [], [],
      ['N'], ['P']))

make('S.Intv.UpdStatus', 'InterviewService', 'UpdateScheduleStatusAsync', 'Service',
    'Verify update schedule status with validation',
    d([('Repository available', [0,1,2])], 'UpdateInterviewStatusDto + Schedule state',
      [('Status=InvalidStatus', [0]), ('Confirm from InProgress', [1]), ('Confirm from Scheduled', [2])],
      [('true', [2])],
      [('ArgumentException', [0]), ('InvalidOperationException', [1])], [],
      ['A','A','N'], ['P','P','P']))

make('S.Intv.JoinRoom', 'InterviewService', 'JoinRoomAsync', 'Service',
    'Verify join room logic with validations',
    d([('Repository available', [0,1,2,3])], 'JoinRoomDto + Room state',
      [('Room not found', [0]), ('Room closed', [1]), ('Room full (1/1)', [2]), ('Room open (0/10)', [3])],
      [('JoinRoomResponseDto', [3])],
      [('KeyNotFoundException', [0]), ('InvalidOperationException("Room da dong")', [1]), ('InvalidOperationException("Room da day")', [2])], [],
      ['A','A','B','N'], ['P','P','P','P']))

make('S.Intv.LeaveRoom', 'InterviewService', 'LeaveRoomAsync', 'Service',
    'Verify leave room logic',
    d([('Repository available', [0,1])], 'LeaveRoomDto + Participant state',
      [('Participant not found', [0]), ('Participant joined', [1])],
      [('false', [0]), ('true (State->Left)', [1])],
      [], [],
      ['A','N'], ['P','P']))

make('S.Intv.Feedback', 'InterviewService', 'SubmitFeedbackAsync', 'Service',
    'Verify submit feedback logic with schedule matching',
    d([('Repository available', [0,1])], 'SubmitFeedbackDto + Assignment state',
      [('Assignment.ScheduleId mismatch', [0]), ('Assignment matches, Pass/8/Good', [1])],
      [('false', [0]), ('true (result+score set)', [1])],
      [], [],
      ['A','N'], ['P','P']))

# Remove original template sheet, rename it
template_ws.title = 'Template (Original)'

wb.save(DST)
print(f'Saved: {DST}')
print(f'Total sheets: {len(wb.sheetnames)}')
ctrl = [s for s in wb.sheetnames if s.startswith('C.')]
svc = [s for s in wb.sheetnames if s.startswith('S.')]
print(f'Controller sheets: {len(ctrl)}')
print(f'Service sheets: {len(svc)}')
